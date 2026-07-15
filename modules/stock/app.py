import os
import socket
import threading
from datetime import datetime
from flask import Flask, render_template, request, redirect, url_for, flash, send_file, current_app, jsonify, abort, session
from sqlalchemy import func, or_
from openpyxl import Workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from extensions import db
from config import Config
from models import (
    Material, Location, Supplier, StockBalance, MaterialMovement, MovementItem,
    ReceiveConfirmation, StockLedger, Attachment, AuditLog
)
from time_utils import now_ist, html_datetime_value
from services import (
    ALL_MOVEMENT_TYPES, TRANSFER_TYPES, RECEIVABLE_TYPES, LOCATION_TYPE_PREFIX,
    movement_label, next_code, create_movement_from_form, receive_movement_from_form,
    audit, save_upload, attach_file, create_backup, list_backups, restore_backup,
    seed_defaults, get_current_user, get_stock, normalize_float, normalize_custom,
    location_has_stock_or_pending
)


def create_app():
    app = Flask(__name__)
    app.config.from_object(Config)
    os.makedirs(app.instance_path, exist_ok=True)
    os.makedirs(app.config["UPLOAD_FOLDER"], exist_ok=True)
    os.makedirs(app.config["EXPORT_FOLDER"], exist_ok=True)
    os.makedirs(app.config["BACKUP_FOLDER"], exist_ok=True)
    db.init_app(app)

    @app.template_filter("movement_label")
    def _movement_label(code):
        return movement_label(code)

    @app.context_processor
    def inject_globals():
        pending_receive_count = 0
        try:
            pending_receive_count = MaterialMovement.query.filter(
                MaterialMovement.status.in_(["In Transit", "Partially Received", "Damaged"]),
                or_(MaterialMovement.movement_type.in_(list(RECEIVABLE_TYPES)),
                    (MaterialMovement.movement_type == "MISC") & (MaterialMovement.stock_effect == "TRANSFER"))
            ).count()
        except Exception:
            pending_receive_count = 0
        def page_url(page):
            args = request.args.to_dict(flat=True)
            args["page"] = page
            return url_for(request.endpoint or "dashboard", **args)
        return {
            "current_user": get_current_user(),
            "now": now_ist(),
            "active_endpoint": request.endpoint or "",
            "pending_receive_count": pending_receive_count,
            "page_url": page_url,
        }


    @app.before_request
    def require_mgroups_portal_stock_access():
        """Protect the stock module when mounted inside the M-GROUPS Portal.

        The main M-GROUPS login writes these session keys:
        admin_id, user_id, user_name, user_role, portal_permissions.
        This module has no standalone login, so every stock URL must be protected here too.
        """
        if request.endpoint == "static" or request.endpoint == "healthz":
            return None
        if not session.get("admin_id"):
            return redirect("/login")

        # Bridge main portal session keys into the inventory app's expected user fields.
        if session.get("admin_id") and not session.get("user_id"):
            session["user_id"] = str(session.get("admin_id"))
        session.setdefault("user_name", session.get("admin_name") or "Portal User")
        session.setdefault("user_role", session.get("admin_role") or "portal_user")

        role = str(session.get("user_role") or "").strip().lower()
        admin_name = str(session.get("admin_name") or session.get("user_name") or "").strip().lower()
        perms = set(session.get("portal_permissions") or [])

        # V15.7.1: legacy admin accounts may still be stored as "admin" or
        # "super_admin" even though the UI displays them as Developer/Owner.
        # They must not be blocked from the stock module.
        full_control_roles = {
            "developer", "owner", "admin", "super_admin", "administrator",
            "developer_owner", "developer_owner_legacy", "developer / owner (legacy)",
        }
        full_control_names = {"admin", "bala", "balachandrakumar"}
        full_control = role in full_control_roles or admin_name in full_control_names or "*" in perms

        def allowed(*keys):
            return full_control or any(k in perms for k in keys)

        if not allowed("stock_view"):
            abort(403)

        path = request.path or "/"

        # Read/export reports need stock_reports.
        if path.startswith("/reports") and not allowed("stock_reports"):
            abort(403)
        if path.startswith("/audit-log") and not allowed("audit_view", "stock_reports"):
            abort(403)

        # Local SQLite backup pages are developer/system only and are not a PostgreSQL backup replacement.
        if path.startswith("/backup") and not allowed("system_settings", "backup_download", "backup_restore"):
            abort(403)

        if request.method == "POST":
            if path.startswith("/materials") or path.startswith("/locations") or path.startswith("/suppliers"):
                if not allowed("stock_add", "stock_adjust"):
                    abort(403)
            elif path.startswith("/receive"):
                if not allowed("stock_add", "stock_transfer"):
                    abort(403)
            elif path.startswith("/movement/create"):
                movement_type = (request.form.get("movement_type") or "").upper()
                if movement_type in {"PURCHASE_TO_STORE", "OPENING_STOCK"} and not allowed("stock_add"):
                    abort(403)
                elif movement_type in {"STORE_TO_SITE", "SITE_TO_SITE", "SITE_TO_STORE_RETURN", "MISC"} and not allowed("stock_transfer", "stock_add"):
                    abort(403)
                elif movement_type in {"STOCK_ADJUSTMENT", "DAMAGE_MISSING", "SITE_CONSUMPTION"} and not allowed("stock_adjust", "stock_add"):
                    abort(403)
                elif not allowed("stock_add", "stock_transfer", "stock_adjust"):
                    abort(403)
        return None

    @app.route("/")
    def dashboard():
        # Include inactive locations if they still have stock, so dashboard never hides real material.
        all_locations = Location.query.order_by(Location.type, Location.name).all()
        materials = Material.query.order_by(Material.name).all()
        stocks = StockBalance.query.all()
        stock_map = {(s.material_id, s.location_id): s.quantity for s in stocks}

        visible_locations = []
        inactive_stock_locations = []
        for loc in all_locations:
            loc_qty = sum((stock_map.get((m.id, loc.id), 0) or 0) for m in materials)
            if loc.status == "Active" or loc_qty > 0:
                visible_locations.append(loc)
            if loc.status != "Active" and loc_qty > 0:
                inactive_stock_locations.append({"location": loc, "qty": loc_qty})

        matrix = []
        low_stock = []
        for mat in materials:
            row = {"material": mat, "locations": {}, "total": 0, "in_transit": 0, "damaged": 0}
            for loc in visible_locations:
                qty = stock_map.get((mat.id, loc.id), 0) or 0
                row["locations"][loc.id] = qty
                row["total"] += qty
            in_transit_items = db.session.query(func.sum(MovementItem.quantity_sent - MovementItem.quantity_received - MovementItem.damaged_quantity)).join(MaterialMovement).filter(
                MovementItem.material_id == mat.id,
                MaterialMovement.status.in_(["In Transit", "Partially Received", "Damaged"]),
                or_(MaterialMovement.movement_type.in_(list(RECEIVABLE_TYPES)),
                    (MaterialMovement.movement_type == "MISC") & (MaterialMovement.stock_effect == "TRANSFER"))
            ).scalar() or 0
            row["in_transit"] = max(0, in_transit_items)
            row["damaged"] = db.session.query(func.sum(MovementItem.damaged_quantity)).filter(MovementItem.material_id == mat.id).scalar() or 0
            if mat.min_stock and row["total"] <= mat.min_stock:
                low_stock.append(row)
            matrix.append(row)

        in_transit_count = MaterialMovement.query.filter_by(status="In Transit").count()
        today_start = now_ist().replace(hour=0, minute=0, second=0, microsecond=0)
        today_movements = MaterialMovement.query.filter(MaterialMovement.created_at >= today_start).count()
        recent = MaterialMovement.query.order_by(MaterialMovement.created_at.desc()).limit(8).all()

        store_value = db.session.query(func.sum(StockBalance.quantity * Material.standard_rate)).join(Material).join(Location).filter(Location.type.in_(["Store", "Godown", "Temporary Yard", "Office"])).scalar() or 0
        site_value = db.session.query(func.sum(StockBalance.quantity * Material.standard_rate)).join(Material).join(Location).filter(Location.type == "Site").scalar() or 0

        material_id = request.args.get("material_id", type=int)
        finder_rows = []
        selected_material = None
        if material_id:
            selected_material = db.session.get(Material, material_id)
            if selected_material:
                for loc in visible_locations:
                    qty = stock_map.get((selected_material.id, loc.id), 0) or 0
                    if qty != 0 or loc.status != "Active":
                        finder_rows.append({"location": loc, "qty": qty})

        return render_template("dashboard.html", locations=visible_locations, materials=materials, matrix=matrix,
                               low_stock=low_stock, in_transit_count=in_transit_count,
                               today_movements=today_movements, recent=recent,
                               store_value=store_value, site_value=site_value,
                               inactive_stock_locations=inactive_stock_locations,
                               selected_material=selected_material, finder_rows=finder_rows)

    # Materials
    @app.route("/materials")
    def materials_list():
        q = request.args.get("q", "").strip()
        page = max(request.args.get("page", 1, type=int) or 1, 1)
        query = Material.query
        if q:
            query = query.filter(Material.name.ilike(f"%{q}%"))
        pagination = query.order_by(Material.name).paginate(page=page, per_page=50, error_out=False)
        return render_template("materials/list.html", materials=pagination.items, q=q, pagination=pagination)

    @app.route("/materials/add", methods=["POST"])
    def materials_add():
        try:
            photo_path = save_upload(request.files.get("photo"), "materials", request.form.get("name", "material"))
            category = normalize_custom(request.form.get("category"), request.form.get("custom_category"))
            unit = normalize_custom(request.form.get("unit"), request.form.get("custom_unit"))
            if not unit:
                raise ValueError("Unit is required.")
            mat = Material(
                material_code=request.form.get("material_code") or next_code("MAT", Material, "material_code"),
                name=request.form["name"],
                category=category,
                sub_category=request.form.get("sub_category", ""),
                brand=request.form.get("brand", ""),
                size_spec=request.form.get("size_spec", ""),
                unit=unit,
                hsn_code=request.form.get("hsn_code", ""),
                gst_percent=normalize_float(request.form.get("gst_percent")),
                min_stock=normalize_float(request.form.get("min_stock")),
                standard_rate=normalize_float(request.form.get("standard_rate")),
                photo_path=photo_path,
                description=request.form.get("description", ""),
                status=request.form.get("status", "Active"),
            )
            db.session.add(mat)
            db.session.flush()
            attach_file("Material", mat.id, photo_path, "Material Photo")
            audit("CREATE", "Material", mat.id, f"Added material {mat.name}")
            db.session.commit()
            flash("Material added.", "success")
        except Exception as exc:
            db.session.rollback()
            flash(str(exc), "danger")
        return redirect(url_for("materials_list"))

    @app.route("/materials/<int:id>/toggle", methods=["POST"])
    def materials_toggle(id):
        mat = db.session.get(Material, id)
        if not mat:
            abort(404)
        stock_total = db.session.query(func.sum(StockBalance.quantity)).filter(StockBalance.material_id == mat.id).scalar() or 0
        if mat.status == "Active" and stock_total > 0:
            flash(f"Cannot deactivate {mat.name}. Stock exists: {stock_total:g} {mat.unit}. Move/consume/adjust stock to zero first.", "danger")
            return redirect(url_for("materials_list"))
        old = mat.status
        mat.status = "Inactive" if mat.status == "Active" else "Active"
        audit("UPDATE", "Material", mat.id, f"Changed status from {old} to {mat.status}", old, mat.status)
        db.session.commit()
        return redirect(url_for("materials_list"))

    # Locations
    @app.route("/locations")
    def locations_list():
        locations = Location.query.order_by(Location.type, Location.name).all()
        stock_totals = {loc.id: (db.session.query(func.sum(StockBalance.quantity)).filter(StockBalance.location_id == loc.id).scalar() or 0) for loc in locations}
        return render_template("locations/list.html", locations=locations, stock_totals=stock_totals)

    @app.route("/locations/add", methods=["POST"])
    def locations_add():
        try:
            loc_type = normalize_custom(request.form.get("type"), request.form.get("custom_type"))
            if not loc_type:
                raise ValueError("Location type is required.")
            prefix = LOCATION_TYPE_PREFIX.get(loc_type, "LOC")
            loc = Location(
                location_code=request.form.get("location_code") or next_code(prefix, Location, "location_code"),
                name=request.form["name"],
                type=loc_type,
                address=request.form.get("address", ""),
                supervisor_name=request.form.get("supervisor_name", ""),
                phone=request.form.get("phone", ""),
                status=request.form.get("status", "Active"),
            )
            db.session.add(loc)
            db.session.flush()
            audit("CREATE", "Location", loc.id, f"Added location {loc.name}")
            db.session.commit()
            flash("Location added.", "success")
        except Exception as exc:
            db.session.rollback()
            flash(str(exc), "danger")
        return redirect(url_for("locations_list"))

    @app.route("/locations/<int:id>/toggle", methods=["POST"])
    def locations_toggle(id):
        loc = db.session.get(Location, id)
        if not loc:
            abort(404)
        if loc.status == "Active":
            stock_total, pending = location_has_stock_or_pending(loc.id)
            if stock_total > 0 or pending > 0:
                flash(f"Cannot deactivate {loc.name}. Current stock: {stock_total:g}. Pending movements: {pending}. Clear these first.", "danger")
                return redirect(url_for("locations_list"))
        old = loc.status
        loc.status = "Inactive" if loc.status == "Active" else "Active"
        audit("UPDATE", "Location", loc.id, f"Changed status from {old} to {loc.status}", old, loc.status)
        db.session.commit()
        return redirect(url_for("locations_list"))

    # Suppliers
    @app.route("/suppliers")
    def suppliers_list():
        suppliers = Supplier.query.order_by(Supplier.name).all()
        return render_template("suppliers/list.html", suppliers=suppliers)

    @app.route("/suppliers/add", methods=["POST"])
    def suppliers_add():
        try:
            sup = Supplier(
                supplier_code=request.form.get("supplier_code") or next_code("SUP", Supplier, "supplier_code"),
                name=request.form["name"],
                contact_person=request.form.get("contact_person", ""),
                phone=request.form.get("phone", ""),
                gst_no=request.form.get("gst_no", ""),
                address=request.form.get("address", ""),
                status=request.form.get("status", "Active"),
            )
            db.session.add(sup)
            db.session.flush()
            audit("CREATE", "Supplier", sup.id, f"Added supplier {sup.name}")
            db.session.commit()
            flash("Supplier added.", "success")
        except Exception as exc:
            db.session.rollback()
            flash(str(exc), "danger")
        return redirect(url_for("suppliers_list"))

    # Movements
    @app.route("/movement")
    def movement_list():
        page = max(request.args.get("page", 1, type=int) or 1, 1)
        pagination = MaterialMovement.query.order_by(MaterialMovement.created_at.desc()).paginate(page=page, per_page=50, error_out=False)
        return render_template("movement/list.html", movements=pagination.items, pagination=pagination)

    @app.route("/movement/new")
    def movement_new():
        selected_type = request.args.get("type", "PURCHASE_TO_STORE")
        return render_template("movement/form.html",
                               movement_types=ALL_MOVEMENT_TYPES,
                               selected_type=selected_type,
                               materials=Material.query.filter_by(status="Active").order_by(Material.name).all(),
                               locations=Location.query.filter_by(status="Active").order_by(Location.type, Location.name).all(),
                               suppliers=Supplier.query.filter_by(status="Active").order_by(Supplier.name).all(),
                               current_datetime=html_datetime_value())

    @app.route("/movement/create", methods=["POST"])
    def movement_create():
        try:
            movement = create_movement_from_form(request.form, request.files)
            db.session.commit()
            flash(f"Movement {movement.movement_no} created.", "success")
            return redirect(url_for("movement_detail", id=movement.id))
        except Exception as exc:
            db.session.rollback()
            flash(str(exc), "danger")
            return redirect(url_for("movement_new", type=request.form.get("movement_type", "PURCHASE_TO_STORE")))

    @app.route("/movement/<int:id>")
    def movement_detail(id):
        movement = db.session.get(MaterialMovement, id)
        if not movement:
            abort(404)
        return render_template("movement/detail.html", movement=movement)

    # Receive
    @app.route("/receive")
    def receive_list():
        movements = MaterialMovement.query.filter(
            MaterialMovement.status.in_(["In Transit", "Partially Received", "Damaged"]),
            or_(MaterialMovement.movement_type.in_(list(RECEIVABLE_TYPES)),
                (MaterialMovement.movement_type == "MISC") & (MaterialMovement.stock_effect == "TRANSFER"))
        ).order_by(MaterialMovement.created_at.desc()).all()
        confirmations = ReceiveConfirmation.query.order_by(ReceiveConfirmation.created_at.desc()).limit(100).all()
        return render_template("receive/list.html", movements=movements, confirmations=confirmations)

    @app.route("/receive/<int:movement_id>")
    def receive_form(movement_id):
        movement = db.session.get(MaterialMovement, movement_id)
        if not movement:
            abort(404)
        return render_template("receive/form.html", movement=movement, current_datetime=html_datetime_value())

    @app.route("/receive/<int:movement_id>/submit", methods=["POST"])
    def receive_submit(movement_id):
        try:
            conf = receive_movement_from_form(movement_id, request.form, request.files)
            db.session.commit()
            flash(f"Receive confirmation saved: {conf.status}.", "success")
            return redirect(url_for("movement_detail", id=movement_id))
        except Exception as exc:
            db.session.rollback()
            flash(str(exc), "danger")
            return redirect(url_for("receive_form", movement_id=movement_id))

    # Stock / Current Stock
    @app.route("/stock")
    def stock_view():
        location_id = request.args.get("location_id", type=int)
        material_id = request.args.get("material_id", type=int)
        q = request.args.get("q", "").strip()
        query = StockBalance.query.join(Material).join(Location)
        if location_id:
            query = query.filter(StockBalance.location_id == location_id)
        if material_id:
            query = query.filter(StockBalance.material_id == material_id)
        if q:
            query = query.filter(Material.name.ilike(f"%{q}%"))
        stocks = query.order_by(Location.name, Material.name).all()

        selected_material = db.session.get(Material, material_id) if material_id else None
        material_locations = []
        if selected_material:
            all_locs = Location.query.order_by(Location.type, Location.name).all()
            for loc in all_locs:
                qty = get_stock(loc.id, selected_material.id)
                if qty != 0 or loc.status != "Active":
                    material_locations.append({"location": loc, "qty": qty})

        return render_template("stock/view.html", stocks=stocks,
                               locations=Location.query.order_by(Location.name).all(),
                               materials=Material.query.order_by(Material.name).all(),
                               sel_location=location_id, sel_material=material_id, q=q,
                               selected_material=selected_material, material_locations=material_locations)

    @app.route("/api/stock/<int:location_id>/<int:material_id>")
    def api_stock(location_id, material_id):
        return jsonify({"quantity": get_stock(location_id, material_id)})

    # Reports
    @app.route("/reports")
    def reports_home():
        rows = db.session.query(Material.name, Material.unit, func.sum(StockBalance.quantity).label("qty"))\
            .join(StockBalance, StockBalance.material_id == Material.id)\
            .group_by(Material.id).order_by(Material.name).all()
        ledgers = StockLedger.query.order_by(StockLedger.created_at.desc()).limit(50).all()
        return render_template("reports/index.html", rows=rows, ledgers=ledgers)

    @app.route("/reports/current-stock/excel")
    def report_current_stock_excel():
        path = os.path.join(current_app.config["EXPORT_FOLDER"], "excel", "current_stock.xlsx")
        os.makedirs(os.path.dirname(path), exist_ok=True)
        wb = Workbook()
        ws = wb.active
        ws.title = "Current Stock"
        ws.append(["Location", "Type", "Location Status", "Material", "Category", "Qty", "Unit", "Rate", "Value", "Updated"])
        for s in StockBalance.query.join(Location).join(Material).order_by(Location.name, Material.name).all():
            ws.append([s.location.name, s.location.type, s.location.status, s.material.name, s.material.category, s.quantity, s.material.unit, s.material.standard_rate, s.quantity * (s.material.standard_rate or 0), s.updated_at.strftime("%Y-%m-%d %H:%M")])
        wb.save(path)
        return send_file(path, as_attachment=True)

    @app.route("/reports/ledger/excel")
    def report_ledger_excel():
        path = os.path.join(current_app.config["EXPORT_FOLDER"], "excel", "stock_ledger.xlsx")
        os.makedirs(os.path.dirname(path), exist_ok=True)
        wb = Workbook()
        ws = wb.active
        ws.title = "Stock Ledger"
        ws.append(["Date", "Ledger No", "Movement No", "Type", "Location", "Material", "Direction", "Qty", "Balance After", "Remarks"])
        for l in StockLedger.query.order_by(StockLedger.created_at.desc()).limit(5000).all():
            ws.append([l.created_at.strftime("%Y-%m-%d %H:%M"), l.ledger_no, l.movement.movement_no if l.movement else "", movement_label(l.movement_type), l.location.name, l.material.name, l.direction, l.quantity, l.balance_after, l.remarks])
        wb.save(path)
        return send_file(path, as_attachment=True)

    @app.route("/reports/movements/excel")
    def report_movements_excel():
        path = os.path.join(current_app.config["EXPORT_FOLDER"], "excel", "material_movements.xlsx")
        os.makedirs(os.path.dirname(path), exist_ok=True)
        wb = Workbook()
        ws = wb.active
        ws.title = "Movements"
        ws.append(["No", "Date", "Type", "Custom Type", "Status", "From", "To", "Supplier", "Vehicle", "Driver", "Material", "Sent", "Received", "Shortage", "Damaged", "Unit", "Rate", "Value", "Entered By", "Remarks"])
        movements = MaterialMovement.query.order_by(MaterialMovement.created_at.desc()).limit(5000).all()
        for m in movements:
            for item in m.items:
                ws.append([m.movement_no, m.movement_datetime.strftime("%Y-%m-%d %H:%M"), movement_label(m.movement_type), m.custom_movement_name, m.status, m.from_location.name if m.from_location else "", m.to_location.name if m.to_location else "", m.supplier.name if m.supplier else "", m.vehicle_no, m.driver_name, item.material.name, item.quantity_sent, item.quantity_received, item.shortage_quantity, item.damaged_quantity, item.unit, item.rate, item.value, m.entered_by, item.remarks])
        wb.save(path)
        return send_file(path, as_attachment=True)

    @app.route("/reports/damage-missing/excel")
    def report_damage_missing_excel():
        path = os.path.join(current_app.config["EXPORT_FOLDER"], "excel", "damage_missing.xlsx")
        os.makedirs(os.path.dirname(path), exist_ok=True)
        wb = Workbook()
        ws = wb.active
        ws.title = "Damage Missing"
        ws.append(["Date", "Movement No", "Location", "Material", "Qty", "Unit", "Type", "Remarks"])
        items = MovementItem.query.join(MaterialMovement).filter(or_(MovementItem.damaged_quantity > 0, MaterialMovement.movement_type == "DAMAGE_MISSING")).order_by(MaterialMovement.created_at.desc()).all()
        for item in items:
            m = item.movement
            qty = item.damaged_quantity or item.quantity_sent
            ws.append([m.movement_datetime.strftime("%Y-%m-%d %H:%M"), m.movement_no, m.from_location.name if m.from_location else (m.to_location.name if m.to_location else ""), item.material.name, qty, item.unit, item.usage_type, item.remarks])
        wb.save(path)
        return send_file(path, as_attachment=True)

    @app.route("/reports/current-stock/pdf")
    def report_current_stock_pdf():
        path = os.path.join(current_app.config["EXPORT_FOLDER"], "pdf", "current_stock.pdf")
        os.makedirs(os.path.dirname(path), exist_ok=True)
        doc = SimpleDocTemplate(path, pagesize=landscape(A4))
        styles = getSampleStyleSheet()
        data = [["Location", "Type", "Status", "Material", "Qty", "Unit", "Value"]]
        for s in StockBalance.query.join(Location).join(Material).order_by(Location.name, Material.name).all():
            data.append([s.location.name, s.location.type, s.location.status, s.material.name, f"{s.quantity:g}", s.material.unit, f"{s.quantity * (s.material.standard_rate or 0):.2f}"])
        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#164635")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ]))
        doc.build([Paragraph("M-Groups Current Stock Report", styles["Title"]), Spacer(1, 12), table])
        return send_file(path, as_attachment=True)

    @app.route("/movement/<int:id>/challan/pdf")
    def movement_challan_pdf(id):
        movement = db.session.get(MaterialMovement, id)
        if not movement:
            abort(404)
        path = os.path.join(current_app.config["EXPORT_FOLDER"], "pdf", f"{movement.movement_no}_challan.pdf")
        os.makedirs(os.path.dirname(path), exist_ok=True)
        doc = SimpleDocTemplate(path, pagesize=A4)
        styles = getSampleStyleSheet()
        data = [["Material", "Qty", "Unit", "Received", "Shortage", "Damaged"]]
        for item in movement.items:
            data.append([item.material.name, f"{item.quantity_sent:g}", item.unit, f"{item.quantity_received:g}", f"{item.shortage_quantity:g}", f"{item.damaged_quantity:g}"])
        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#164635")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ]))
        parts = [
            Paragraph("M-Groups Material Movement Challan", styles["Title"]), Spacer(1, 12),
            Paragraph(f"Movement No: {movement.movement_no}", styles["Normal"]),
            Paragraph(f"Type: {movement_label(movement.movement_type)} {movement.custom_movement_name or ''}", styles["Normal"]),
            Paragraph(f"From: {movement.from_location.name if movement.from_location else '-'}", styles["Normal"]),
            Paragraph(f"To: {movement.to_location.name if movement.to_location else '-'}", styles["Normal"]),
            Paragraph(f"Vehicle / Driver: {movement.vehicle_no or '-'} / {movement.driver_name or '-'}", styles["Normal"]),
            Spacer(1, 12), table, Spacer(1, 24),
            Paragraph("Sent By: ____________________    Driver: ____________________    Received By: ____________________", styles["Normal"]),
        ]
        doc.build(parts)
        return send_file(path, as_attachment=True)

    # Audit
    @app.route("/audit-log")
    def audit_log():
        page = max(request.args.get("page", 1, type=int) or 1, 1)
        pagination = AuditLog.query.order_by(AuditLog.created_at.desc()).paginate(page=page, per_page=50, error_out=False)
        return render_template("audit/log.html", logs=pagination.items, pagination=pagination)

    # Backup
    @app.route("/backup")
    def backup_home():
        return render_template("backup/index.html", backups=list_backups())

    @app.route("/backup/create", methods=["POST"])
    def backup_create():
        try:
            filename = create_backup()
            db.session.commit()
            flash(f"Backup created: {filename}", "success")
        except Exception as exc:
            db.session.rollback()
            flash(str(exc), "danger")
        return redirect(url_for("backup_home"))

    @app.route("/backup/download/<path:filename>")
    def backup_download(filename):
        path = os.path.join(current_app.config["BACKUP_FOLDER"], os.path.basename(filename))
        if not os.path.exists(path):
            flash("Backup file not found.", "danger")
            return redirect(url_for("backup_home"))
        return send_file(path, as_attachment=True)

    @app.route("/backup/restore", methods=["POST"])
    def backup_restore():
        filename = request.form.get("filename", "")
        confirm = request.form.get("confirm", "")
        try:
            if confirm != "YES":
                raise ValueError("Type YES to restore a backup.")
            safety = restore_backup(filename)
            db.session.remove()
            audit("RESTORE", "Backup", None, f"Restored backup {filename}; safety backup {safety}")
            db.session.commit()
            # The Windows runner restarts automatically when Flask exits with code 3.
            threading.Timer(1.2, lambda: os._exit(3)).start()
            return render_template("backup/restarting.html", filename=filename, safety=safety)
        except Exception as exc:
            flash(str(exc), "danger")
        return redirect(url_for("backup_home"))

    @app.route("/healthz")
    def healthz():
        return {"ok": True, "app": "M-Groups Inventory V6.2.2 Portal Stock"}

    with app.app_context():
        db.create_all()
        seed_defaults()

    return app


app = create_app()

if __name__ == "__main__":
    host = os.environ.get("MGROUPS_HOST", "0.0.0.0")
    port = int(os.environ.get("MGROUPS_PORT", "5000"))
    debug = os.environ.get("MGROUPS_DEBUG", "0") == "1"
    local_url = f"http://127.0.0.1:{port}"
    try:
        sock = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        sock.connect(("8.8.8.8", 80))
        lan_ip = sock.getsockname()[0]
        sock.close()
    except Exception:
        lan_ip = "YOUR-PC-IP"
    print("=" * 62)
    print("M-Groups Inventory V6.2.1")
    print(f"Local URL : {local_url}")
    print(f"LAN URL   : http://{lan_ip}:{port}")
    print("Phone access works only on the same WiFi/LAN and if Windows Firewall allows port 5000.")
    print("For site access outside office WiFi, deploy to cloud/VPS with PostgreSQL.")
    print("=" * 62)
    app.run(host=host, port=port, debug=debug, use_reloader=False)
