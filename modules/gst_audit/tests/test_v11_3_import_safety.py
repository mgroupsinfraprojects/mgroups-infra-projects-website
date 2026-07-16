from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from app.core.import_safety import analyze_import_set, export_import_safety_report, period_label


def _add_b2b_sheet(wb: Workbook, title: str = "B2B"):
    ws = wb.create_sheet(title) if title not in wb.sheetnames else wb[title]
    ws.append(["Goods and Services Tax  - GSTR-2B"])
    ws.append([])
    ws.append([])
    ws.append(["Taxable inward supplies received from registered persons"])
    ws.append([
        "GSTIN of supplier", "Trade/Legal name", "Invoice Details", "", "", "",
        "Place of supply", "Supply Attract Reverse Charge", "Taxable Value (₹)", "Tax Amount", "", "",
        "GSTR-1/1A/IFF/GSTR-5 Period", "GSTR-1/1A/IFF/GSTR-5 Filing Date", "ITC Availability",
    ])
    ws.append(["", "", "Invoice number", "Invoice type", "Invoice Date", "Invoice Value(₹)", "", "", "", "Integrated Tax(₹)", "Central Tax(₹)", "State/UT Tax(₹)", "", "", ""])
    ws.append(["33ABCDE1234F1Z5", "ALPHA TRADERS", "A-001", "Regular", "01/04/2025", 1180, "Tamil Nadu", "No", 1000, 0, 90, 90, "Apr'25", "10/05/2025", "Yes"])
    ws.append(["33ABCDE1234F1Z5", "ALPHA TRADERS", "A-002", "Regular", "02/04/2025", 590, "Tamil Nadu", "No", 500, 0, 45, 45, "Apr'25", "10/05/2025", "Yes"])
    return ws


def _make_b2b_only(path: Path):
    wb = Workbook()
    wb.active.title = "B2B"
    _add_b2b_sheet(wb, "B2B")
    wb.save(path)


def _make_full_workbook(path: Path):
    wb = Workbook()
    wb.active.title = "Read me"
    wb.active.append(["Tax Period", "Apr'25"])
    wb.create_sheet("ITC Available")
    _add_b2b_sheet(wb, "B2B")
    wb.create_sheet("B2BA")
    wb.create_sheet("B2B-CDNR")
    wb.save(path)


def test_import_safety_prefers_full_workbook_and_excludes_b2b_duplicate(tmp_path):
    full = tmp_path / "APRIL 25.xlsx"
    b2b = tmp_path / "4 APRIL 25.xlsx"
    _make_full_workbook(full)
    _make_b2b_only(b2b)

    report = analyze_import_set([str(b2b), str(full)])

    assert report.status == "SAFE_WITH_DUPLICATES_EXCLUDED"
    assert report.uploaded_file_count == 2
    assert report.unique_period_count == 1
    assert report.selected_file_count == 1
    assert report.duplicate_file_count == 1
    assert report.recommended_file_names() == ["APRIL 25.xlsx"]
    assert report.excluded_file_names() == ["4 APRIL 25.xlsx"]
    assert report.duplicate_groups[0].exact_duplicate is True
    assert report.duplicate_groups[0].prevented_rows == 2
    assert report.score >= 88


def test_import_safety_blocks_conflicting_period_duplicates(tmp_path):
    full = tmp_path / "APRIL 25.xlsx"
    b2b = tmp_path / "4 APRIL 25.xlsx"
    _make_full_workbook(full)
    _make_b2b_only(b2b)
    # Change one invoice value so the period duplicates are not identical.
    from openpyxl import load_workbook
    wb = load_workbook(b2b)
    ws = wb["B2B"]
    ws["F7"] = 9999
    wb.save(b2b)

    report = analyze_import_set([str(b2b), str(full)])

    assert report.blocked is True
    assert report.status == "BLOCKED_DUPLICATES_NEED_MANUAL_REVIEW"
    assert report.unresolved_duplicate_periods == ("2025-04",)
    assert report.duplicate_groups[0].exact_duplicate is False


def test_import_safety_report_export(tmp_path):
    full = tmp_path / "APRIL 25.xlsx"
    b2b = tmp_path / "4 APRIL 25.xlsx"
    out = tmp_path / "import_safety.xlsx"
    _make_full_workbook(full)
    _make_b2b_only(b2b)
    report = analyze_import_set([str(b2b), str(full)])

    export_import_safety_report(report, out)

    assert out.exists()
    assert out.stat().st_size > 1000
    assert period_label("2025-04") == "Apr 2025"
