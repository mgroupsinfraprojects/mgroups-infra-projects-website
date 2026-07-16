"""v9.4 regression and correctness tests.

Covers:
  B-NEW-1  Credit note guard must trigger only on invoice_value < 0
  B-NEW-2  Exporter mismatch sheet must exclude CREDIT_NOTE_BALANCED / ZERO_RATED
  D-2      unique_suppliers / unique_gstins must not count excluded official rows
"""
from __future__ import annotations

from datetime import date
from decimal import Decimal
from pathlib import Path
from unittest.mock import MagicMock

import pandas as pd
import pytest
from openpyxl import Workbook

from app.core.audit_engine import (
    InvoiceAuditEngine,
    classify_gst_mismatch_details,
    is_gst_invoice_detail_sheet,
)
from app.core.config import AuditConfig
from app.core.gstin import calculate_gstin_checksum
from app.core.models import AuditResult, AuditSummary, InvoiceRow


# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------

def valid_gstin(prefix: str = "33ABCDE1234F1Z") -> str:
    return prefix + calculate_gstin_checksum(prefix)


def _row(row_id: int, *, invoice_no: str = "INV-1", value: str = "118.00",
         sheet: str = "B2B", gstin: str | None = None, include: bool = True) -> InvoiceRow:
    g = gstin or valid_gstin()
    r = InvoiceRow(
        row_id=row_id,
        source_file="sample.xlsx",
        sheet_name=sheet,
        excel_row_number=row_id,
        raw_snapshot=["raw"],
        supplier_name="ABC Traders",
        gstin=g,
        invoice_no=invoice_no,
        invoice_date=date(2026, 1, 1),
        invoice_series="INV",
        invoice_sequence_no=row_id,
        taxable_value=Decimal("100.00"),
        cgst=Decimal("9.00"),
        sgst=Decimal("9.00"),
        invoice_value=Decimal(value),
        expected_invoice_value=Decimal("118.00"),
        difference_amount=Decimal(value) - Decimal("118.00"),
        mismatch_reason="BALANCED_OR_ROUNDING",
        audit_status="VALID" if include else "DUPLICATE_EXCLUDED",
        include_in_totals=include,
        review_required=False,
        review_decision="ACCEPTED_AUTO",
        duplicate_key=f"{g}|{invoice_no}|2026-01-01",
    )
    r.detected_snapshot = {"gstin": r.gstin}
    r.final_snapshot = dict(r.detected_snapshot)
    return r


# ---------------------------------------------------------------------------
# B-NEW-1  Credit note guard
# ---------------------------------------------------------------------------

class TestCreditNoteGuard:
    def test_negative_invoice_all_negative_is_balanced_credit_note(self):
        d = classify_gst_mismatch_details(
            Decimal("-118.00"), Decimal("-118.00"), Decimal("-100.00"), Decimal("-18.00"),
        )
        assert d["reason"] == "CREDIT_NOTE_BALANCED"
        assert d["review"] is False
        assert d["include"] is True

    def test_negative_invoice_zero_rated_credit_note(self):
        d = classify_gst_mismatch_details(
            Decimal("-50.00"), Decimal("0"), Decimal("0"), Decimal("0"),
        )
        assert d["reason"] == "CREDIT_NOTE_ZERO_RATED"
        assert d["include"] is True

    def test_negative_invoice_unbalanced_is_credit_note_mismatch(self):
        d = classify_gst_mismatch_details(
            Decimal("-200.00"), Decimal("-118.00"), Decimal("-100.00"), Decimal("-18.00"),
        )
        assert d["reason"] == "CREDIT_NOTE_MISMATCH"
        assert d["review"] is True
        assert d["include"] is False

    def test_positive_invoice_negative_taxable_is_component_sign_mismatch(self):
        """v9.3 regression: positive invoice must NOT enter the credit note branch."""
        d = classify_gst_mismatch_details(
            Decimal("118.00"), Decimal("118.00"), Decimal("-100.00"), Decimal("218.00"),
        )
        assert d["reason"] == "COMPONENT_SIGN_MISMATCH"
        assert d["review"] is True
        assert d["include"] is False

    def test_positive_invoice_negative_gst_total_is_component_sign_mismatch(self):
        d = classify_gst_mismatch_details(
            Decimal("100.00"), Decimal("100.00"), Decimal("100.00"), Decimal("-18.00"),
        )
        assert d["reason"] == "COMPONENT_SIGN_MISMATCH"
        assert d["review"] is True

    def test_normal_valid_invoice_not_affected(self):
        d = classify_gst_mismatch_details(
            Decimal("118.00"), Decimal("118.00"), Decimal("100.00"), Decimal("18.00"),
        )
        assert d["reason"] == "BALANCED_OR_ROUNDING"
        assert d["include"] is True


# ---------------------------------------------------------------------------
# B-NEW-2  Exporter mismatch sheet exclusion
# ---------------------------------------------------------------------------

class TestExporterMismatchSheet:
    def _make_result(self) -> AuditResult:
        rows = []
        # normal valid row
        r1 = _row(1, invoice_no="INV-1")
        r1.mismatch_reason = "BALANCED_OR_ROUNDING"
        rows.append(r1)
        # balanced credit note — must NOT appear on mismatch tab
        r2 = _row(2, invoice_no="CN-1", value="-118.00")
        r2.mismatch_reason = "CREDIT_NOTE_BALANCED"
        rows.append(r2)
        # zero-rated credit note — must NOT appear on mismatch tab
        r3 = _row(3, invoice_no="CN-2", value="-5.00")
        r3.mismatch_reason = "CREDIT_NOTE_ZERO_RATED"
        rows.append(r3)
        # genuine mismatch — MUST appear on mismatch tab
        r4 = _row(4, invoice_no="INV-4", value="999.00", include=False)
        r4.mismatch_reason = "UNEXPLAINED_GST_MISMATCH"
        r4.review_required = True
        r4.audit_status = "REVIEW_REQUIRED"
        rows.append(r4)

        summary = AuditSummary(files_processed=1, sheets_processed=1, classified_rows=4)
        return AuditResult(rows=rows, summary=summary, source_totals={}, month_totals={}, supplier_totals={})

    def test_credit_note_balanced_excluded_from_mismatch_sheet(self, tmp_path):
        from app.core.exporter import export_verified_excel
        result = self._make_result()
        out = tmp_path / "report.xlsx"
        export_verified_excel(result, str(out))

        df = pd.read_excel(out, sheet_name="Mismatch Reasons", engine="openpyxl", header=1)
        reasons = df["mismatch_reason"].tolist() if "mismatch_reason" in df.columns else []
        assert "CREDIT_NOTE_BALANCED" not in reasons, \
            "CREDIT_NOTE_BALANCED must not appear on the Mismatch Reasons sheet"
        assert "CREDIT_NOTE_ZERO_RATED" not in reasons, \
            "CREDIT_NOTE_ZERO_RATED must not appear on the Mismatch Reasons sheet"
        assert "UNEXPLAINED_GST_MISMATCH" in reasons, \
            "Genuine mismatches must still appear"


# ---------------------------------------------------------------------------
# D-2  unique_suppliers must only count approved rows
# ---------------------------------------------------------------------------

class TestSupplierCountAccuracy:
    def test_excluded_official_rows_not_counted_in_unique_suppliers(self, tmp_path):
        """Duplicate-excluded rows from B2B sheet must not inflate supplier count."""
        path = tmp_path / "test.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "B2B"
        # Header
        ws.append(["GSTIN of supplier", "Trade/Legal name", "Invoice number",
                   "Invoice date", "Taxable Value", "CGST", "SGST", "IGST", "Invoice Value"])
        g1 = valid_gstin("33ABCDE1234F1Z")
        g2 = valid_gstin("27XYZPQ5678G2Z")
        # supplier 1 — two identical invoices (one will be DUPLICATE_EXCLUDED)
        ws.append([g1, "ABC Traders", "INV-1", "01-01-2026", 100, 9, 9, 0, 118])
        ws.append([g1, "ABC Traders", "INV-1", "01-01-2026", 100, 9, 9, 0, 118])
        # supplier 2 — unique invoice
        ws.append([g2, "XYZ Suppliers", "INV-2", "01-01-2026", 200, 18, 18, 0, 236])
        wb.save(path)

        result = InvoiceAuditEngine(AuditConfig()).process_files([str(path)])

        # One duplicate row should be excluded
        assert result.summary.duplicate_rows == 1
        # Only 2 unique suppliers with approved rows — not 3
        assert result.summary.unique_suppliers == 2, (
            f"Expected 2 unique suppliers, got {result.summary.unique_suppliers}. "
            "Excluded rows must not inflate supplier count."
        )
        assert result.summary.unique_gstins == 2

    def test_unique_suppliers_counts_all_approved_when_no_official_rows(self, tmp_path):
        """Fallback path (non-B2B sheet) still counts approved rows correctly."""
        path = tmp_path / "test.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"  # not a B2B/CDNR tab
        ws.append(["GSTIN of supplier", "Trade/Legal name", "Invoice number",
                   "Invoice date", "Taxable Value", "CGST", "SGST", "IGST", "Invoice Value"])
        g1 = valid_gstin("33ABCDE1234F1Z")
        g2 = valid_gstin("27XYZPQ5678G2Z")
        ws.append([g1, "ABC Traders", "INV-A", "01-01-2026", 100, 9, 9, 0, 118])
        ws.append([g2, "XYZ Suppliers", "INV-B", "01-01-2026", 200, 18, 18, 0, 236])
        wb.save(path)

        result = InvoiceAuditEngine(AuditConfig()).process_files([str(path)])
        assert result.summary.unique_suppliers == 2
        assert result.summary.unique_gstins == 2


# ---------------------------------------------------------------------------
# M-1  _flag_uncertain_detection helper exists and is callable
# ---------------------------------------------------------------------------

def test_flag_uncertain_detection_helper_deduplicated():
    """Ensure the helper method exists and sets the correct flags atomically."""
    engine = InvoiceAuditEngine()
    row = _row(1)
    row.audit_status = "VALID"

    detection = MagicMock()
    detection.uncertain = True
    detection.warning = "Header detection was uncertain."

    engine._flag_uncertain_detection([row], detection)

    assert row.review_required is True
    assert row.include_in_totals is False
    assert row.audit_status == "REVIEW_REQUIRED"
    assert "Header detection was uncertain." in row.audit_notes


def test_flag_uncertain_detection_skips_already_skipped_rows():
    engine = InvoiceAuditEngine()
    row = _row(1)
    row.audit_status = "SKIPPED_EMPTY"
    row.review_required = False
    row.include_in_totals = False

    detection = MagicMock()
    detection.uncertain = True
    detection.warning = "uncertain"

    engine._flag_uncertain_detection([row], detection)
    # Should not change a SKIPPED_EMPTY row
    assert row.audit_status == "SKIPPED_EMPTY"
    assert row.review_required is False


# ---------------------------------------------------------------------------
# v9.5 completion: supplier count semantics + standard export performance contract
# ---------------------------------------------------------------------------

def test_supplier_count_exposes_detected_and_approved_semantics(tmp_path):
    path = tmp_path / "test.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "B2B"
    ws.append(["GSTIN of supplier", "Trade/Legal name", "Invoice number",
               "Invoice date", "Taxable Value", "CGST", "SGST", "IGST", "Invoice Value"])
    g1 = valid_gstin("33ABCDE1234F1Z")
    g2 = valid_gstin("27XYZPQ5678G2Z")
    ws.append([g1, "ABC Traders", "INV-1", "01-01-2026", 100, 9, 9, 0, 118])
    ws.append([g1, "ABC Traders", "INV-1", "01-01-2026", 100, 9, 9, 0, 118])
    ws.append([g2, "XYZ Suppliers", "INV-2", "01-01-2026", 200, 18, 18, 0, 236])
    wb.save(path)

    result = InvoiceAuditEngine(AuditConfig()).process_files([str(path)])
    assert result.summary.unique_suppliers == 2
    assert result.summary.unique_gstins == 2
    assert result.summary.detected_unique_suppliers == 2
    assert result.summary.detected_unique_gstins == 2


def test_standard_export_omits_heavy_forensic_snapshot_columns(tmp_path):
    from app.core.exporter import export_verified_excel
    result = InvoiceAuditEngine(AuditConfig()).build_result_from_rows([_row(1)], 1, 1)
    out = tmp_path / "standard.xlsx"
    export_verified_excel(result, str(out))
    df = pd.read_excel(out, sheet_name="All Classified Rows", engine="openpyxl", header=1)
    assert "raw_snapshot" not in df.columns
    assert "detected_snapshot" not in df.columns
    assert "final_snapshot" not in df.columns
    assert "invoice_value" in df.columns


def test_forensic_export_can_include_snapshot_columns(tmp_path):
    from app.core.exporter import export_verified_excel
    result = InvoiceAuditEngine(AuditConfig()).build_result_from_rows([_row(1)], 1, 1)
    out = tmp_path / "forensic.xlsx"
    export_verified_excel(result, str(out), include_forensic_columns=True)
    df = pd.read_excel(out, sheet_name="All Classified Rows", engine="openpyxl", header=1)
    assert "raw_snapshot" in df.columns
    assert "detected_snapshot" in df.columns
    assert "final_snapshot" in df.columns
