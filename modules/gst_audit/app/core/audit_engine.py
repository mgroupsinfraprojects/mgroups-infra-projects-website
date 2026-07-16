from __future__ import annotations

import csv
import logging
import re
from collections import defaultdict
from decimal import Decimal
from pathlib import Path
from statistics import median
from typing import Callable, Dict, Iterable, Iterator, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook

from app.core.config import AuditConfig
from app.core.date_parser import parse_invoice_date
from app.core.field_detector import FieldDetector
from app.core.gstin import detect_gstin_roles, find_all_gstins, find_gstin, is_valid_gstin, normalize_text
from app.core.hsn import HSN_SAC_PATTERN, extract_hsn_sac, validate_hsn_sac
from app.core.invoice_number import parse_invoice_number
from app.core.models import AuditResult, AuditSummary, InvoiceRow
from app.core.money import to_decimal
from app.core.performance import ProcessingStats

LOGGER = logging.getLogger(__name__)
ProgressCallback = Callable[[int, str], None]


class UnsupportedFileTypeError(ValueError):
    """Raised when upload preflight rejects a file type before parsing."""


SUPPORTED_EXTENSIONS = {".csv", ".tsv", ".xlsx", ".xlsm", ".xls"}

LOW_ROUNDING_LIMIT = Decimal("1.00")
MINOR_LIMIT = Decimal("5.00")
NON_SUPPLIER_TOKENS = frozenset({
    "regular",
    "tamil nadu",
    "composition",
    "unregistered",
    "yes",
    "no",
})

GST_PORTAL_HEADER_TOKENS = frozenset({
    "invoice type",
    "invoice date",
    "invoice value",
    "place of supply",
    "supply attract reverse charge",
    "taxable value",
    "integrated tax",
    "central tax",
    "state/ut tax",
    "cess",
    "gstr-1",
    "gstr-5",
    "gstr-6",
    "itc availability",
    "applicable % of tax rate",
    "irn date",
    "note number",
    "note type",
    "note value",
    "document number",
    "document type",
    "bill of entry",
    "icegate reference",
})

GST_PORTAL_SUPPORT_TOKENS = frozenset({
    "credit which may be availed",
    "credit which may not be availed",
    "credit which is rejected",
    "inward supplies from isd",
    "all other itc",
    "supplies from registered persons",
    "itc reversal",
    "form gstr-3b",
    "part a",
    "part b",
    "details b2b",
    "others",
})
SUPPLIER_SUFFIX_TOKENS = (
    "PRIVATE LIMITED", "PVT LTD", "PVT. LTD.", "LIMITED", "LTD", "LLP",
    "ENTERPRISES", "ENTERPRISE", "TRADERS", "TRADING", "AGENCIES", "AGENCY",
    "INDUSTRIES", "INDUSTRY", "CORPORATION", "CO", "COMPANY",
)


def _mask_gstin(value: object) -> str:
    text = normalize_text(value).upper().replace(" ", "")
    if len(text) <= 6:
        return "***" if text else ""
    return f"{text[:2]}{'*' * max(0, len(text) - 6)}{text[-4:]}"


def supplier_group_key(row: InvoiceRow) -> str:
    """Use GSTIN as the stable supplier aggregation key; display name is secondary."""
    if row.gstin:
        display = row.supplier_name or "UNKNOWN"
        return f"{row.gstin} — {display}"
    return row.supplier_name or "UNKNOWN"


def is_gst_invoice_detail_sheet(sheet_name: str) -> bool:
    """Return True for GST portal detail tabs that contain source invoice/note rows.

    This deliberately excludes summary/readme/ITC roll-up tabs so the dashboard can
    separate physical rows scanned from actual GST invoice/detail rows.
    """
    name = re.sub(r"[^a-z0-9]", "", str(sheet_name or "").lower())
    if not name:
        return False
    return name.startswith(("b2b", "eco", "isd", "impg"))


def is_official_invoice_detail_row(row: InvoiceRow) -> bool:
    if row.audit_status.startswith("ERROR") or row.audit_status in {
        "SKIPPED_EMPTY",
        "SKIPPED_HEADER_OR_TITLE",
        "SKIPPED_TOTAL_OR_NON_INVOICE_ROW",
    }:
        return False
    return is_gst_invoice_detail_sheet(row.sheet_name) and bool(row.gstin) and bool(row.invoice_no)


def _safe_cell(value: object) -> str:
    return normalize_text(value)


def _row_values(series_or_values: Iterable[object]) -> List[str]:
    return [_safe_cell(v) for v in list(series_or_values)]


def _is_effectively_empty(values: Iterable[str]) -> bool:
    return all(not str(v).strip() for v in values)


def _is_header_like(values: Iterable[str]) -> bool:
    cells = [normalize_text(v).strip() for v in values]
    text = " ".join(str(v).lower() for v in cells if str(v).strip())
    if not text:
        return False
    header_hits = [
        "goods and services tax",
        "taxable inward supplies",
        "gstin of supplier",
        "gstin of recipient",
        "trade/legal name",
        "invoice details",
        "invoice number",
        "tax amount",
        "return period",
        "filing date",
    ]
    if any(hit in text for hit in header_hits):
        return True

    token_hits = sum(1 for token in GST_PORTAL_HEADER_TOKENS if token in text)
    support_hits = sum(1 for token in GST_PORTAL_SUPPORT_TOKENS if token in text)
    has_gstin = bool(find_gstin(cells))
    money_cells = [to_decimal(v) for v in cells]
    has_money = any(abs(value) > Decimal("0.00") for value in money_cells)

    # GSTR-2B exports contain many section/header bands such as
    # "Invoice type / Invoice date / Invoice value / Place of supply".
    # They must stay out of the audit review queue. Real invoice rows normally
    # have a GSTIN and/or amount; header bands have several known column labels
    # and no business identity.
    if not has_gstin and not has_money and token_hits >= 2:
        return True
    if not has_gstin and support_hits >= 1 and token_hits >= 1:
        return True
    return False


def _looks_like_gst_portal_support_row(values: List[str], sheet_name: str = "") -> bool:
    text = " ".join(str(v or "").lower() for v in values)
    sheet = str(sheet_name or "").lower()
    has_gstin = bool(find_gstin(values))
    has_money = any(abs(to_decimal(v)) > Decimal("0.00") for v in values)
    header_hits = sum(1 for token in GST_PORTAL_HEADER_TOKENS if token in text)
    support_hits = sum(1 for token in GST_PORTAL_SUPPORT_TOKENS if token in text or token in sheet)
    if not has_gstin and not has_money and header_hits >= 2:
        return True
    if not has_gstin and support_hits >= 1 and header_hits >= 1:
        return True
    if not has_gstin and support_hits >= 2:
        return True
    return False


def _get_by_map(values: List[str], field_map: Dict[str, int], field: str) -> str:
    col = field_map.get(field)
    if col is None or col >= len(values):
        return ""
    return values[col]


def _first_text_candidate(values: List[str], exclude_gstin: str = "") -> str:
    for text in values:
        t = normalize_text(text)
        if not t:
            continue
        if exclude_gstin and exclude_gstin in t.upper():
            continue
        if len(t) >= 3 and not t.replace(".", "", 1).isdigit():
            lower = t.lower()
            if lower in NON_SUPPLIER_TOKENS:
                continue
            return t
    return ""


def clean_supplier_name(value: str) -> str:
    """Normalize supplier names for display/search without destroying raw data."""
    text = re.sub(r"\s+", " ", normalize_text(value)).strip(" -_,.;")
    changed = True
    while changed and text:
        changed = False
        upper = text.upper().rstrip(" .,")
        for suffix in SUPPLIER_SUFFIX_TOKENS:
            pattern = r"(?:\s+|^)" + re.escape(suffix) + r"\.?$"
            new_text = re.sub(pattern, "", upper, flags=re.IGNORECASE).strip(" -_,.;")
            if new_text != upper and len(new_text) >= 3:
                text = new_text.title()
                changed = True
                break
    return text or normalize_text(value)


def detect_hsn_sac(values: List[str], field_map: Dict[str, int]) -> str:
    mapped = _get_by_map(values, field_map, "hsn_sac")
    if mapped:
        code = extract_hsn_sac(mapped)
        if code:
            return code
    # Do not blindly scan every numeric token: invoice numbers like INV-00001
    # otherwise become false HSN values. Without an HSN/SAC mapped column, only
    # accept cells that explicitly mention HSN/SAC.
    if "hsn_sac" not in field_map:
        for value in values:
            text = str(value or "")
            if "hsn" in text.lower() or "sac" in text.lower():
                code = extract_hsn_sac(text)
                if code:
                    return code
    return ""


def classify_gst_mismatch_details(
    invoice_value: Decimal,
    expected_value: Decimal,
    taxable: Decimal,
    gst_total: Decimal,
    *,
    config: AuditConfig | None = None,
) -> Dict[str, object]:
    cfg = config or AuditConfig()
    difference = invoice_value - expected_value
    diff = difference.copy_abs()
    suggestion = ""
    # A credit or debit note is defined by a negative note value (invoice_value < 0).
    # Records with a positive invoice value but a negative taxable/GST component are
    # data quality issues (wrong sign on an adjustment line) and must NOT be treated
    # as credit notes — doing so would silently flip include_in_totals to False for
    # normal invoices with any negative component.
    if invoice_value < 0:
        if taxable == 0 and gst_total == 0:
            return {
                "reason": "CREDIT_NOTE_ZERO_RATED",
                "severity": "LOW",
                "review": False,
                "include": True,
                "suggestion": "Zero-rated/exempt credit or debit note with no taxable/GST components.",
            }
        expected_neg = taxable + gst_total
        credit_diff = (invoice_value - expected_neg).copy_abs()
        if credit_diff <= cfg.low_rounding_limit:
            return {
                "reason": "CREDIT_NOTE_BALANCED",
                "severity": "LOW",
                "review": False,
                "include": True,
                "suggestion": "Balanced credit/debit note.",
            }
        return {
            "reason": "CREDIT_NOTE_MISMATCH",
            "severity": "HIGH",
            "review": True,
            "include": False,
            "suggestion": f"Credit/debit note: expected {expected_neg}, got {invoice_value}, diff {invoice_value - expected_neg}.",
        }
    # Positive invoice value with a negative taxable or GST component is a data
    # quality problem (sign error on an adjustment line), not a credit note.
    if taxable < 0 or gst_total < 0:
        return {
            "reason": "COMPONENT_SIGN_MISMATCH",
            "severity": "HIGH",
            "review": True,
            "include": False,
            "suggestion": (
                f"Invoice value {invoice_value} is positive but taxable={taxable} or "
                f"gst_total={gst_total} is negative. Check for a sign error on an "
                f"adjustment line or wrong column mapping."
            ),
        }
    if invoice_value == 0 and expected_value == 0:
        return {"reason": "NO_AMOUNT_DETECTED", "severity": "LOW", "review": False, "include": False, "suggestion": "Skipped as empty, header, summary, or non-invoice row. It remains traceable but does not need manual review."}
    if diff <= cfg.low_rounding_limit:
        return {"reason": "BALANCED_OR_ROUNDING", "severity": "LOW", "review": False, "include": True, "suggestion": "No correction needed; difference is within rounding tolerance."}
    if diff <= cfg.minor_rounding_limit:
        return {"reason": "MINOR_ROUNDING_OR_DECIMAL_ISSUE", "severity": "LOW", "review": False, "include": True, "suggestion": f"Acceptable small difference of {difference}; verify decimal rounding if needed."}
    small_percent = (diff / max(invoice_value.copy_abs(), Decimal("1.00")) * Decimal("100.00")) if invoice_value != 0 else Decimal("100.00")
    if diff <= cfg.auto_accept_small_difference_limit and small_percent <= cfg.auto_accept_small_percent_limit:
        return {"reason": "AUTO_ACCEPT_SMALL_DIFFERENCE", "severity": "LOW", "review": False, "include": True, "suggestion": f"Small non-material difference of {difference} auto accepted under configured threshold."}
    if diff <= cfg.mandatory_review_amount_limit and small_percent <= cfg.auto_accept_small_percent_limit and invoice_value > 0:
        return {"reason": "AUTO_ACCEPT_SMALL_EXPENSE_OR_ROUNDING", "severity": "LOW", "review": False, "include": True, "suggestion": f"Small non-material percentage difference of {difference} auto accepted; retained in trace log."}
    if taxable > 0 and gst_total == 0:
        return {"reason": "GST_COMPONENT_MISSING_OR_EXEMPT_CHECK", "severity": "HIGH", "review": True, "include": False, "suggestion": "Taxable value exists but GST components are zero. Check exempt/zero-rated supply or shifted tax columns."}
    if invoice_value < taxable:
        return {"reason": "INVOICE_LESS_THAN_TAXABLE_POSSIBLE_COLUMN_SHIFT", "severity": "HIGH", "review": True, "include": False, "suggestion": "Invoice value is lower than taxable value. Check column shift or wrong amount mapping."}
    if diff in {Decimal("500.00"), Decimal("1000.00")}:
        suggestion = f"Taxable + GST = {expected_value}; invoice shows {invoice_value}. Difference {difference} may be freight, discount, round-off, TCS/TDS, or extra charge."
        return {"reason": "POSSIBLE_FREIGHT_OR_DISCOUNT", "severity": "MEDIUM", "review": True, "include": False, "suggestion": suggestion}
    if invoice_value > expected_value and diff <= cfg.max_mismatch_reasonable_expense:
        reason = "INVOICE_VALUE_INCLUDES_EXPENSES"
        if diff % Decimal("1.00") != 0:
            reason = "ROUNDING_IN_MULTIPLE_COMPONENTS"
        suggestion = f"Expected {expected_value}, actual {invoice_value}, difference {difference}. Verify freight, packing, round-off, reimbursement, TCS/TDS, or expense columns."
        return {"reason": reason, "severity": "MEDIUM", "review": True, "include": False, "suggestion": suggestion}
    if invoice_value > 0 and (diff / max(invoice_value.copy_abs(), Decimal("1.00"))) < Decimal("0.01"):
        return {"reason": "SMALL_PERCENTAGE_DIFFERENCE_CHECK_FREIGHT_ROUNDING", "severity": "MEDIUM", "review": True, "include": False, "suggestion": f"Difference is below 1% ({difference}). Review for freight/round-off or extra charge."}
    if diff > Decimal("0.00") and gst_total > 0 and (diff / max(gst_total.copy_abs(), Decimal("1.00"))) < Decimal("0.10"):
        return {"reason": "TCS_TDS_DETECTED", "severity": "MEDIUM", "review": True, "include": False, "suggestion": f"Difference {difference} is small relative to GST total; check TCS/TDS or withholding lines."}
    return {"reason": "UNEXPLAINED_GST_MISMATCH", "severity": "HIGH", "review": True, "include": False, "suggestion": f"Expected invoice value is {expected_value}; actual invoice value is {invoice_value}; difference is {difference}. Manual review required."}


def classify_gst_mismatch(
    invoice_value: Decimal,
    expected_value: Decimal,
    taxable: Decimal,
    gst_total: Decimal,
) -> Tuple[str, str, bool, bool]:
    # Backwards-compatible compact classifier used by old tests/scripts.
    # The v5 engine uses classify_gst_mismatch_details() for richer reasons.
    detail = classify_gst_mismatch_details(invoice_value, expected_value, taxable, gst_total)
    reason = str(detail["reason"])
    severity = str(detail["severity"])
    if reason in {"INVOICE_VALUE_INCLUDES_EXPENSES", "ROUNDING_IN_MULTIPLE_COMPONENTS"}:
        reason = "UNEXPLAINED_GST_MISMATCH"
        severity = "HIGH"
    return reason, severity, bool(detail["review"]), bool(detail["include"])


class InvoiceAuditEngine:
    """Non-destructive GST invoice audit processor with v5 scale/intelligence features."""

    def __init__(self, config: AuditConfig | None = None) -> None:
        self.config = config or AuditConfig()
        self.detector = FieldDetector()

    def process_files(
        self,
        file_paths: List[str],
        progress_callback: Optional[ProgressCallback] = None,
        ignored_gstins: Optional[Iterable[str]] = None,
        self_gstins: Optional[Iterable[str]] = None,
        config: AuditConfig | None = None,
    ) -> AuditResult:
        cfg = (config or self.config).with_runtime_values(ignored_gstins=ignored_gstins, self_gstins=self_gstins)
        rows: List[InvoiceRow] = []
        row_id = 1
        sheets_processed = 0
        total_files = len(file_paths)
        stats = ProcessingStats.start()
        from app.version import APP_VERSION
        LOGGER.info("Starting GST audit v%s processing for %s file(s)", APP_VERSION, total_files)

        for file_index, file_path in enumerate(file_paths, start=1):
            path = Path(file_path)
            self._progress(progress_callback, int((file_index - 1) / max(total_files, 1) * 100), stats.progress_message(f"Opening {path.name}"))
            try:
                self._preflight_file(path, cfg)
                if path.suffix.lower() in {".csv", ".tsv"}:
                    produced, row_id = self._process_csv(path, row_id, rows, cfg, stats, progress_callback, file_index, total_files)
                    sheets_processed += produced
                elif path.suffix.lower() in {".xlsx", ".xlsm"}:
                    produced, row_id = self._process_excel_streaming(path, row_id, rows, cfg, stats, progress_callback, file_index, total_files)
                    sheets_processed += produced
                else:
                    produced, row_id = self._process_excel_pandas(path, row_id, rows, cfg, stats, progress_callback, file_index, total_files)
                    sheets_processed += produced
            except Exception as exc:
                LOGGER.exception("Failed to process file %s", path)
                message = str(exc)
                status = "ERROR_FILE_UNREADABLE"
                severity = "CRITICAL"
                if isinstance(exc, UnsupportedFileTypeError):
                    status = "ERROR_UNSUPPORTED_FILE_TYPE"
                    severity = "MEDIUM"
                elif "xlrd" in message.lower():
                    status = "ERROR_MISSING_DEPENDENCY"
                    severity = "HIGH"
                rows.append(self._make_skipped_row(
                    row_id=row_id,
                    source_file=path.name,
                    sheet_name="FILE_ERROR",
                    excel_row_number=0,
                    raw_snapshot=[message],
                    status=status,
                    notes=f"File was isolated from totals: {message}",
                    severity=severity,
                ))
                row_id += 1

        self._mark_duplicates(rows)
        if cfg.enable_invoice_gap_detection:
            self._annotate_invoice_gaps(rows)
        if cfg.enable_supplier_anomaly_detection:
            self._annotate_supplier_anomalies(rows, cfg)
        result = self.build_result_from_rows(rows, total_files, sheets_processed)
        self._progress(progress_callback, 100, stats.progress_message(f"Completed: {result.summary.final_status}"))
        LOGGER.info("Finished GST audit v%s: rows=%s approved=%s review=%s status=%s", APP_VERSION, result.summary.raw_rows_read, result.summary.final_approved_rows, result.summary.review_required_rows, result.summary.final_status)
        return result

    def _preflight_file(self, path: Path, cfg: AuditConfig) -> None:
        if not path.exists():
            raise FileNotFoundError(f"File not found: {path}")
        suffix = path.suffix.lower()
        if suffix not in SUPPORTED_EXTENSIONS:
            if suffix == ".pdf":
                raise UnsupportedFileTypeError("PDF import is not supported in this version. Upload GST portal XLSX/XLS/CSV exports instead.")
            raise UnsupportedFileTypeError(f"Unsupported file type '{suffix or '<none>'}'. Supported formats: XLSX, XLSM, XLS, CSV, TSV.")
        size_mb = path.stat().st_size / (1024 * 1024)
        if size_mb > cfg.max_file_size_mb:
            raise ValueError(f"File exceeds safety limit: {size_mb:.1f} MB > {cfg.max_file_size_mb} MB")

    def _enforce_sheet_limits(self, *, sheet_name: str, rows_seen: int, column_count: int, cfg: AuditConfig) -> None:
        if rows_seen > cfg.max_rows_per_file:
            raise ValueError(f"Row limit exceeded in {sheet_name}: {rows_seen} > {cfg.max_rows_per_file}")
        if column_count > cfg.max_columns_per_sheet:
            raise ValueError(f"Column limit exceeded in {sheet_name}: {column_count} > {cfg.max_columns_per_sheet}")

    @staticmethod
    def _flag_uncertain_detection(rows: List[InvoiceRow], detection: "FieldDetectionResult") -> None:
        """Mark the last appended row as REVIEW_REQUIRED when header detection was uncertain.

        Extracted from three processing paths (Excel streaming, Excel pandas, CSV) to
        eliminate code duplication and ensure consistent behaviour across all formats.
        """
        if not rows:
            return
        last = rows[-1]
        if detection.uncertain and last.audit_status not in {"SKIPPED_EMPTY", "SKIPPED_HEADER_OR_TITLE"}:
            last.review_required = True
            last.include_in_totals = False
            last.audit_status = "REVIEW_REQUIRED"
            last.audit_severity = "HIGH"
            last.audit_indicator = "⚠️"
            last.review_decision = "PENDING_REVIEW"
            last.append_audit_note(detection.warning)

    def _process_excel_streaming(self, path: Path, row_id: int, rows: List[InvoiceRow], cfg: AuditConfig, stats: ProcessingStats, progress_callback: Optional[ProgressCallback], file_index: int, total_files: int) -> Tuple[int, int]:
        workbook = load_workbook(path, read_only=True, data_only=True)
        if len(workbook.worksheets) > cfg.max_sheets_per_file:
            workbook.close()
            raise ValueError(f"Sheet limit exceeded: {len(workbook.worksheets)} > {cfg.max_sheets_per_file}")
        sheets_processed = 0
        try:
            for sheet_index, worksheet in enumerate(workbook.worksheets, start=1):
                sheets_processed += 1
                iterator = worksheet.iter_rows(values_only=True)
                preview: List[List[object]] = []
                for _ in range(cfg.header_scan_rows):
                    try:
                        preview.append(list(next(iterator)))
                    except StopIteration:
                        break
                    if len(preview) % 50 == 0:
                        self._progress(progress_callback, min(int(file_index / max(total_files, 1) * 100), 99), stats.progress_message(f"Scanning header {path.name} / {worksheet.title}"))

                frame = pd.DataFrame(preview)
                detection = self.detector.detect(frame, max_scan_rows=cfg.header_scan_rows, min_score=cfg.header_min_score)
                if detection.uncertain:
                    LOGGER.warning("Header detection uncertain: file=%s sheet=%s score=%s", path.name, worksheet.title, detection.confidence_score)

                def combined_rows() -> Iterator[List[object]]:
                    for row in preview:
                        yield row
                    for row in iterator:
                        yield list(row)

                stream = iter(combined_rows())
                try:
                    current = next(stream)
                except StopIteration:
                    continue

                zero_idx = 0
                for nxt in stream:
                    values = _row_values(current)
                    next_cache = _row_values(nxt)
                    self._enforce_sheet_limits(sheet_name=worksheet.title, rows_seen=zero_idx + 1, column_count=len(values), cfg=cfg)
                    row_id = self._process_values_row(row_id, rows, path.name, worksheet.title, zero_idx + 1, zero_idx, values, detection.field_map, detection.data_start, next_cache, cfg)
                    self._flag_uncertain_detection(rows, detection)
                    stats.rows_processed += 1
                    if stats.rows_processed % 50 == 0:
                        pct = int(((file_index - 1) + sheet_index / max(len(workbook.worksheets), 1)) / max(total_files, 1) * 100)
                        self._progress(progress_callback, min(pct, 99), stats.progress_message(f"Processing {path.name} / {worksheet.title}"))
                    current = nxt
                    zero_idx += 1

                values = _row_values(current)
                self._enforce_sheet_limits(sheet_name=worksheet.title, rows_seen=zero_idx + 1, column_count=len(values), cfg=cfg)
                row_id = self._process_values_row(row_id, rows, path.name, worksheet.title, zero_idx + 1, zero_idx, values, detection.field_map, detection.data_start, [], cfg)
                self._flag_uncertain_detection(rows, detection)
                stats.rows_processed += 1
                pct = int(((file_index - 1) + sheet_index / max(len(workbook.worksheets), 1)) / max(total_files, 1) * 100)
                self._progress(progress_callback, min(pct, 99), stats.progress_message(f"Processed {path.name} / {worksheet.title}"))
        finally:
            workbook.close()
        return sheets_processed, row_id

    def _process_excel_pandas(self, path: Path, row_id: int, rows: List[InvoiceRow], cfg: AuditConfig, stats: ProcessingStats, progress_callback: Optional[ProgressCallback], file_index: int, total_files: int) -> Tuple[int, int]:
        excel = pd.ExcelFile(path)
        if len(excel.sheet_names) > cfg.max_sheets_per_file:
            raise ValueError(f"Sheet limit exceeded: {len(excel.sheet_names)} > {cfg.max_sheets_per_file}")
        sheets_processed = 0
        for sheet_index, sheet_name in enumerate(excel.sheet_names, start=1):
            sheets_processed += 1
            df = pd.read_excel(path, sheet_name=sheet_name, header=None, dtype=object)
            detection = self.detector.detect(df, max_scan_rows=cfg.header_scan_rows, min_score=cfg.header_min_score)
            if detection.uncertain:
                LOGGER.warning("Header detection uncertain: file=%s sheet=%s score=%s", path.name, sheet_name, detection.confidence_score)
            for zero_idx in range(len(df)):
                values = _row_values(df.iloc[zero_idx])
                next_values = _row_values(df.iloc[zero_idx + 1]) if zero_idx + 1 < len(df) else []
                self._enforce_sheet_limits(sheet_name=sheet_name, rows_seen=zero_idx + 1, column_count=len(values), cfg=cfg)
                row_id = self._process_values_row(row_id, rows, path.name, sheet_name, zero_idx + 1, zero_idx, values, detection.field_map, detection.data_start, next_values, cfg)
                if detection.uncertain and rows[-1].audit_status not in {"SKIPPED_EMPTY", "SKIPPED_HEADER_OR_TITLE"}:
                    rows[-1].review_required = True
                    rows[-1].include_in_totals = False
                    rows[-1].audit_status = "REVIEW_REQUIRED"
                    rows[-1].audit_severity = "HIGH"
                    rows[-1].audit_indicator = "⚠️"
                    rows[-1].review_decision = "PENDING_REVIEW"
                    rows[-1].audit_notes = (rows[-1].audit_notes + "; " if rows[-1].audit_notes else "") + detection.warning
                stats.rows_processed += 1
                if stats.rows_processed % 50 == 0:
                    pct = int(((file_index - 1) + sheet_index / max(len(excel.sheet_names), 1)) / max(total_files, 1) * 100)
                    self._progress(progress_callback, min(pct, 99), stats.progress_message(f"Processing {path.name} / {sheet_name}"))
            pct = int(((file_index - 1) + sheet_index / max(len(excel.sheet_names), 1)) / max(total_files, 1) * 100)
            self._progress(progress_callback, min(pct, 99), stats.progress_message(f"Processed {path.name} / {sheet_name}"))
        return sheets_processed, row_id

    def _detect_csv_encoding(self, path: Path) -> str:
        """Detect common GST/accounting CSV encodings without silently corrupting text."""
        raw = path.read_bytes()[:8192]
        if raw.startswith(b"\xff\xfe"):
            return "utf-16"
        if raw.startswith(b"\xfe\xff"):
            return "utf-16"
        if raw.startswith(b"\xef\xbb\xbf"):
            return "utf-8-sig"
        candidates = ("utf-8-sig", "utf-16", "utf-16le", "utf-16be", "cp1252", "iso-8859-1")
        best_encoding = "utf-8-sig"
        best_score = -1.0
        for encoding in candidates:
            try:
                text = raw.decode(encoding)
            except UnicodeError:
                continue
            if not text:
                continue
            printable = sum(1 for ch in text if ch.isprintable() or ch in "\r\n\t")
            replacement_penalty = text.count("�") * 5
            nul_penalty = text.count("\x00") * 10
            score = (printable - replacement_penalty - nul_penalty) / max(len(text), 1)
            if score > best_score:
                best_score = score
                best_encoding = encoding
        if best_score < 0.60:
            LOGGER.warning(
                "CSV encoding detection weak for %s. Selected %s with score %.2f; output may require manual review.",
                path.name, best_encoding, best_score,
            )
        try:
            with path.open("r", encoding=best_encoding, newline="") as fh:
                fh.read(2048)
        except UnicodeError:
            LOGGER.warning(
                "CSV encoding detection failed for %s after trying UTF-8/UTF-16/Windows encodings. "
                "Defaulting to utf-8-sig; output may contain garbled characters.",
                path.name,
            )
            return "utf-8-sig"
        return best_encoding

    def _detect_delimiter(self, path: Path, encoding: str) -> str:
        """Detect delimiters used by GST portal, Tally, Busy, Zoho, and ERP CSV exports."""
        if path.suffix.lower() == ".tsv":
            return "\t"
        try:
            with path.open("r", encoding=encoding, newline="", errors="replace") as fh:
                sample = fh.read(8192)
            if sample.lower().startswith("sep=") and len(sample) >= 5:
                return sample[4]
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|^")
                return dialect.delimiter
            except Exception:
                pass
            lines = [line for line in sample.splitlines() if line.strip() and not line.lower().startswith("sep=")]
            scores: dict[str, tuple[int, int]] = {}
            for delimiter in (",", ";", "\t", "|", "^"):
                counts = [line.count(delimiter) for line in lines[:40]]
                non_zero = [count for count in counts if count > 0]
                if non_zero:
                    # Prefer high column count with stable counts across rows.
                    scores[delimiter] = (sum(non_zero), -len(set(non_zero)))
            if scores:
                return max(scores.items(), key=lambda item: item[1])[0]
        except Exception as exc:
            LOGGER.warning("CSV delimiter detection failed for %s: %s", path.name, exc)
        return ","

    def _process_csv(self, path: Path, row_id: int, rows: List[InvoiceRow], cfg: AuditConfig, stats: ProcessingStats, progress_callback: Optional[ProgressCallback], file_index: int, total_files: int) -> Tuple[int, int]:
        """Stream GST/accounting CSV/TSV files with metadata rows and uneven row widths.

        v9.6 keeps only the header preview in memory, then processes the file with
        a one-row lookahead. This preserves cancellation/progress behavior and
        avoids the v9.5 all-rows CSV buffer on large 100k+/500k-row exports.
        """
        encoding = self._detect_csv_encoding(path)
        delimiter = self._detect_delimiter(path, encoding)
        preview_limit = max(cfg.header_scan_rows, 100)
        preview_rows: List[tuple[int, List[object]]] = []
        max_columns_seen = 0

        with path.open("r", encoding=encoding, newline="", errors="replace") as fh:
            reader = csv.reader(fh, delimiter=delimiter)
            physical_row = 0
            for physical_row, raw in enumerate(reader, start=1):
                row_list = list(raw)
                preview_rows.append((physical_row, row_list))
                max_columns_seen = max(max_columns_seen, len(row_list))
                if physical_row >= preview_limit:
                    break

            if not preview_rows:
                rows.append(self._make_skipped_row(row_id, path.name, "CSV", 0, [], "SKIPPED_EMPTY", "CSV file is empty. Not used in totals."))
                return 1, row_id + 1

            self._enforce_sheet_limits(sheet_name="CSV", rows_seen=len(preview_rows), column_count=max_columns_seen, cfg=cfg)
            detection_frame = pd.DataFrame([raw for _row_no, raw in preview_rows])
            detection = self.detector.detect(detection_frame, max_scan_rows=preview_limit, min_score=cfg.header_min_score)
            if detection.uncertain:
                LOGGER.warning("Header detection uncertain: file=%s sheet=CSV score=%s", path.name, detection.confidence_score)

            def remaining_records() -> Iterator[tuple[int, List[object]]]:
                nonlocal physical_row
                for row_no, raw_values in preview_rows:
                    yield row_no, raw_values
                for raw_values in reader:
                    physical_row += 1
                    yield physical_row, list(raw_values)

            sheet_name = "CSV"
            records = iter(remaining_records())
            current = next(records, None)
            zero_idx = 0
            while current is not None:
                next_record = next(records, None)
                physical_row_no, raw_values = current
                values = _row_values(raw_values)
                next_values = _row_values(next_record[1]) if next_record is not None else []
                max_columns_seen = max(max_columns_seen, len(values))
                if physical_row_no % 50 == 0:
                    self._enforce_sheet_limits(sheet_name="CSV", rows_seen=physical_row_no, column_count=max_columns_seen, cfg=cfg)
                row_id = self._process_values_row(row_id, rows, path.name, sheet_name, physical_row_no, zero_idx, values, detection.field_map, detection.data_start, next_values, cfg)
                self._flag_uncertain_detection(rows, detection)
                stats.rows_processed += 1
                if stats.rows_processed % 50 == 0:
                    self._progress(progress_callback, min(int(file_index / max(total_files, 1) * 100), 99), stats.progress_message(f"Processing CSV {path.name}"))
                zero_idx += 1
                current = next_record

        self._enforce_sheet_limits(sheet_name="CSV", rows_seen=zero_idx, column_count=max_columns_seen, cfg=cfg)
        self._progress(progress_callback, min(int(file_index / max(total_files, 1) * 100), 99), stats.progress_message(f"Processed CSV {path.name}"))
        return 1, row_id

    def _process_values_row(self, row_id: int, rows: List[InvoiceRow], source_file: str, sheet_name: str, excel_row_number: int, zero_idx: int, values: List[str], field_map: Dict[str, int], data_start: int, next_values: List[str], cfg: AuditConfig) -> int:
        if _is_effectively_empty(values):
            rows.append(self._make_skipped_row(row_id, source_file, sheet_name, excel_row_number, values, "SKIPPED_EMPTY", "Empty physical row. Not used in totals."))
            return row_id + 1
        first_cell = values[0].strip().lower() if values else ""
        if first_cell.startswith("sep="):
            rows.append(self._make_skipped_row(row_id, source_file, sheet_name, excel_row_number, values, "SKIPPED_HEADER_OR_TITLE", "CSV delimiter declaration row. Not used in totals."))
            return row_id + 1
        if zero_idx < data_start or _is_header_like(values):
            rows.append(self._make_skipped_row(row_id, source_file, sheet_name, excel_row_number, values, "SKIPPED_HEADER_OR_TITLE", "Header/title/metadata row. Not used in totals."))
            return row_id + 1
        rows.append(self._parse_invoice_like_row(row_id, source_file, sheet_name, excel_row_number, values, field_map, next_values, cfg))
        return row_id + 1

    def build_result_from_rows(self, rows: List[InvoiceRow], files_processed: int = 0, sheets_processed: int = 0) -> AuditResult:
        summary = self.build_summary(rows, files_processed, sheets_processed)
        return AuditResult(rows=rows, summary=summary, source_totals=self.group_totals(rows, "source_file"), month_totals=self.month_totals(rows), supplier_totals=self.supplier_totals(rows))

    def recalculate_result(self, result: AuditResult) -> AuditResult:
        # Manual review decisions can change include_in_totals. Re-run the same
        # post-processing checks that are required after initial ingestion so
        # duplicate, gap, and anomaly flags cannot silently disappear.
        for row in result.rows:
            if row.audit_status == "DUPLICATE_EXCLUDED":
                row.audit_status = "REVIEW_REQUIRED"
                row.audit_severity = "HIGH"
                row.audit_indicator = "⚠️"
                row.review_required = True
                row.review_decision = "PENDING_REVIEW"
                row.include_in_totals = False
        self._mark_duplicates(result.rows)
        if self.config.enable_invoice_gap_detection:
            self._annotate_invoice_gaps(result.rows)
        if self.config.enable_supplier_anomaly_detection:
            self._annotate_supplier_anomalies(result.rows, self.config)
        rebuilt = self.build_result_from_rows(result.rows, result.summary.files_processed, result.summary.sheets_processed)
        result.summary = rebuilt.summary
        result.source_totals = rebuilt.source_totals
        result.month_totals = rebuilt.month_totals
        result.supplier_totals = rebuilt.supplier_totals
        return result

    @staticmethod
    def _progress(callback: Optional[ProgressCallback], percent: int, message: str) -> None:
        if callback:
            callback(max(0, min(100, int(percent))), message)

    @staticmethod
    def _make_skipped_row(row_id: int, source_file: str, sheet_name: str, excel_row_number: int, raw_snapshot: List[str], status: str, notes: str, severity: str = "LOW") -> InvoiceRow:
        return InvoiceRow(row_id=row_id, source_file=source_file, sheet_name=sheet_name, excel_row_number=excel_row_number, raw_snapshot=raw_snapshot, audit_status=status, audit_severity=severity, audit_indicator="❌", audit_notes=notes, include_in_totals=False, review_required=False, review_decision="EXCLUDED")

    def _parse_invoice_like_row(self, row_id: int, source_file: str, sheet_name: str, excel_row_number: int, values: List[str], field_map: Dict[str, int], next_values: List[str], cfg: AuditConfig) -> InvoiceRow:
        roles = detect_gstin_roles(values, self_gstins=cfg.self_gstins)
        gstin = roles.supplier_gstin or find_gstin(values) or ""
        supplier = clean_supplier_name(_get_by_map(values, field_map, "supplier_name") or _first_text_candidate(values, gstin))
        invoice_no = _get_by_map(values, field_map, "invoice_no")
        hsn_sac = detect_hsn_sac(values, field_map)
        hsn_result = validate_hsn_sac(hsn_sac)
        invoice_info = parse_invoice_number(invoice_no)
        invoice_date_value = _get_by_map(values, field_map, "invoice_date")
        invoice_date, date_status = parse_invoice_date(invoice_date_value)
        period = _get_by_map(values, field_map, "period")

        taxable = to_decimal(_get_by_map(values, field_map, "taxable_value"))
        igst = to_decimal(_get_by_map(values, field_map, "igst"))
        cgst = to_decimal(_get_by_map(values, field_map, "cgst"))
        sgst = to_decimal(_get_by_map(values, field_map, "sgst"))
        cess = to_decimal(_get_by_map(values, field_map, "cess"))
        invoice_value = to_decimal(_get_by_map(values, field_map, "invoice_value"))

        if _looks_like_gst_portal_support_row(values, sheet_name):
            row = self._make_skipped_row(
                row_id,
                source_file,
                sheet_name,
                excel_row_number,
                values,
                "SKIPPED_HEADER_OR_SUPPORT_ROW",
                "GST portal header/ITC/support row detected. Not an invoice row and not sent for approval.",
            )
            row.taxable_value = taxable
            row.igst = igst
            row.cgst = cgst
            row.sgst = sgst
            row.cess = cess
            row.invoice_value = Decimal("0.00")
            row.detected_snapshot = {"reason": "gst portal header/support row"}
            row.final_snapshot = dict(row.detected_snapshot)
            return row

        reconstructed = False
        reconstruction_notes: List[str] = []
        if next_values and not find_gstin(next_values):
            next_text_candidate = _first_text_candidate(next_values, gstin)
            if (
                supplier
                and next_text_candidate
                and next_text_candidate not in supplier
                and len(next_text_candidate) >= 8
                and not next_text_candidate.replace(",", "").replace(".", "").isdigit()
                and not any(kw in next_text_candidate.lower() for kw in ("total", "sub-total", "subtotal", "grand", "amount"))
            ):
                supplier = clean_supplier_name(f"{supplier} {next_text_candidate}".strip())
                reconstructed = True
                reconstruction_notes.append("Supplier continuation text found in next row.")
            if not invoice_no:
                numeric_or_text = next((v for v in next_values if v.strip()), "")
                if numeric_or_text:
                    invoice_no = numeric_or_text
                    invoice_info = parse_invoice_number(invoice_no)
                    reconstructed = True
                    reconstruction_notes.append("Invoice number recovered from continuation row candidate.")

        if not gstin and not is_valid_gstin(gstin, require_checksum=True) and not any(abs(v) > Decimal("0.00") for v in [taxable, igst, cgst, sgst, cess, invoice_value]):
            row = self._make_skipped_row(row_id, source_file, sheet_name, excel_row_number, values, "SKIPPED_HEADER_OR_SUPPORT_ROW", "No GSTIN and no amount detected. Treated as header/support row, not an invoice approval item.")
            row.supplier_name = supplier
            row.invoice_no = invoice_no
            row.detected_snapshot = {"reason": "no gstin and no amount"}
            row.final_snapshot = dict(row.detected_snapshot)
            return row

        if not gstin and not supplier and not invoice_no:
            row = self._make_skipped_row(row_id, source_file, sheet_name, excel_row_number, values, "SKIPPED_TOTAL_OR_NON_INVOICE_ROW", "No GSTIN/supplier/invoice number detected. Treated as total/metadata/non-invoice row; not used in dashboard totals.")
            row.taxable_value = taxable
            row.igst = igst
            row.cgst = cgst
            row.sgst = sgst
            row.cess = cess
            row.invoice_value = Decimal("0.00")
            row.detected_snapshot = {"reason": "no invoice identity detected"}
            row.final_snapshot = dict(row.detected_snapshot)
            return row

        if gstin and gstin.upper() in cfg.ignored_gstins:
            row = self._make_skipped_row(row_id, source_file, sheet_name, excel_row_number, values, "IGNORED_GSTIN_EXCLUDED", "GSTIN is configured in ignored GSTIN list. Row is excluded from approved dashboard totals.")
            row.supplier_name = supplier
            row.gstin = gstin
            row.recipient_gstin = roles.recipient_gstin
            row.all_gstins = roles.all_gstins
            row.gstin_roles_note = roles.note
            row.self_invoice_flag = roles.self_invoice
            row.invoice_no = invoice_no
            row.hsn_sac = hsn_sac
            row.hsn_valid = hsn_result.is_valid
            row.hsn_notes = hsn_result.note
            row.taxable_value = taxable
            row.igst = igst
            row.cgst = cgst
            row.sgst = sgst
            row.cess = cess
            row.invoice_value = invoice_value
            row.detected_snapshot = {"gstin": gstin, "supplier_name": supplier, "invoice_no": invoice_no, "hsn_sac": hsn_sac, "reason": "ignored_gstin"}
            row.final_snapshot = dict(row.detected_snapshot)
            return row

        gst_total = igst + cgst + sgst + cess
        expected = taxable + gst_total
        difference = invoice_value - expected
        diff_abs = difference.copy_abs()
        diff_pct = Decimal("0.00") if invoice_value == 0 else (diff_abs / max(invoice_value.copy_abs(), Decimal("1.00"))) * Decimal("100.00")
        detail = classify_gst_mismatch_details(invoice_value, expected, taxable, gst_total, config=cfg)
        mismatch_reason, severity, review_required, include = str(detail["reason"]), str(detail["severity"]), bool(detail["review"]), bool(detail["include"])
        suggested_correction = str(detail.get("suggestion", ""))

        notes: List[str] = []
        if roles.note:
            notes.append(roles.note)
        if roles.self_invoice and cfg.classify_self_invoices_as_review:
            notes.append("Supplier GSTIN equals configured self GSTIN; review possible self-invoice/self-entry.")
            severity = "HIGH"
            review_required = True
            include = False
        if not gstin:
            notes.append("GSTIN not detected.")
            severity = "HIGH"
            review_required = True
            include = False
        elif not is_valid_gstin(gstin, require_checksum=True):
            notes.append("GSTIN checksum/format invalid.")
            severity = "HIGH"
            review_required = True
            include = False
            LOGGER.warning("GSTIN validation failed: file=%s sheet=%s row=%s gstin=%s", source_file, sheet_name, excel_row_number, _mask_gstin(gstin))
        if not supplier:
            notes.append("Supplier name missing or not detected.")
            severity = "MEDIUM" if severity == "LOW" else severity
            review_required = True
            include = False
        if not invoice_no:
            notes.append("Invoice number missing or not detected.")
            severity = "MEDIUM" if severity == "LOW" else severity
            review_required = True
            include = False
        if date_status not in {"DATE_OK", "DATE_OK_FALLBACK"}:
            notes.append(date_status)
            severity = "MEDIUM" if severity == "LOW" else severity
            review_required = True
            include = False
        if hsn_sac and not hsn_result.is_valid:
            notes.append(hsn_result.note)
            severity = "MEDIUM" if severity == "LOW" else severity
            review_required = True
            include = False
        if mismatch_reason not in {"BALANCED_OR_ROUNDING", "MINOR_ROUNDING_OR_DECIMAL_ISSUE", "CREDIT_NOTE_BALANCED", "CREDIT_NOTE_ZERO_RATED"}:
            notes.append(mismatch_reason)
            if suggested_correction:
                notes.append(f"Suggestion: {suggested_correction}")
        if reconstructed:
            notes.extend(reconstruction_notes)
            review_required = True
            include = False
            severity = "MEDIUM" if severity == "LOW" else severity
            LOGGER.info("Row reconstruction flagged: file=%s sheet=%s row=%s notes=%s", source_file, sheet_name, excel_row_number, "; ".join(reconstruction_notes))

        if include:
            status = "VALID" if not reconstructed else "WARNING_RECONSTRUCTED"
            indicator = "✅" if status == "VALID" else "🔧"
            decision = "ACCEPTED_AUTO"
        elif review_required:
            status = "REVIEW_REQUIRED"
            indicator = "⚠️" if not reconstructed else "🔧"
            decision = "PENDING_REVIEW"
        else:
            status = "EXCLUDED"
            indicator = "❌"
            decision = "EXCLUDED"

        duplicate_key = f"{gstin.upper()}|{invoice_no.strip().upper()}|{invoice_date.isoformat() if invoice_date else ''}"
        row = InvoiceRow(
            row_id=row_id,
            source_file=source_file,
            sheet_name=sheet_name,
            excel_row_number=excel_row_number,
            raw_snapshot=values,
            supplier_name=supplier,
            gstin=gstin,
            recipient_gstin=roles.recipient_gstin,
            all_gstins=roles.all_gstins,
            self_invoice_flag=roles.self_invoice,
            gstin_roles_note=roles.note,
            invoice_no=invoice_no,
            invoice_series=invoice_info.series,
            invoice_sequence_no=invoice_info.sequence,
            hsn_sac=hsn_sac,
            hsn_valid=hsn_result.is_valid,
            hsn_notes=hsn_result.note,
            invoice_date=invoice_date,
            period=period,
            taxable_value=taxable,
            igst=igst,
            cgst=cgst,
            sgst=sgst,
            cess=cess,
            invoice_value=invoice_value,
            expected_invoice_value=expected,
            difference_amount=difference,
            difference_percent=diff_pct.quantize(Decimal("0.01")),
            mismatch_reason=mismatch_reason,
            suggested_correction=suggested_correction,
            audit_status=status,
            audit_severity=severity,
            audit_indicator=indicator,
            audit_notes="; ".join(notes) if notes else "Balanced and accepted automatically.",
            review_required=review_required,
            review_decision=decision,
            include_in_totals=include,
            reconstructed=reconstructed,
            duplicate_key=duplicate_key,
        )
        row.detected_snapshot = {
            "gstin": gstin,
            "recipient_gstin": roles.recipient_gstin,
            "all_gstins": roles.all_gstins,
            "self_invoice_flag": roles.self_invoice,
            "supplier_name": supplier,
            "invoice_no": invoice_no,
            "invoice_series": invoice_info.series,
            "invoice_sequence_no": invoice_info.sequence,
            "hsn_sac": hsn_sac,
            "hsn_valid": hsn_result.is_valid,
            "invoice_date": invoice_date.isoformat() if invoice_date else "",
            "invoice_value": float(invoice_value),
            "expected_invoice_value": float(expected),
            "difference_amount": float(difference),
            "suggested_correction": suggested_correction,
        }
        row.final_snapshot = dict(row.detected_snapshot)
        return row

    def _mark_duplicates(self, rows: List[InvoiceRow]) -> None:
        seen: Dict[str, int] = {}
        for row in rows:
            if row.audit_status.startswith("SKIPPED") or row.audit_status.startswith("ERROR") or not row.gstin or not row.invoice_no:
                continue
            if row.duplicate_key in seen:
                row.audit_status = "DUPLICATE_EXCLUDED"
                row.audit_severity = "HIGH"
                row.audit_indicator = "🔁"
                row.append_audit_note(f"Duplicate of row id {seen[row.duplicate_key]}. Excluded from totals.")
                row.review_required = True
                row.review_decision = "DUPLICATE_PENDING_REVIEW"
                row.include_in_totals = False
                LOGGER.warning("Duplicate invoice excluded: file=%s sheet=%s row=%s duplicate_of=%s gstin=%s", row.source_file, row.sheet_name, row.excel_row_number, seen[row.duplicate_key], _mask_gstin(row.gstin))
            else:
                seen[row.duplicate_key] = row.row_id

    def _annotate_invoice_gaps(self, rows: List[InvoiceRow]) -> None:
        groups: Dict[tuple[str, str], List[InvoiceRow]] = defaultdict(list)
        for row in rows:
            if row.gstin and row.invoice_series and row.invoice_sequence_no is not None and not row.audit_status.startswith("SKIPPED"):
                groups[(row.gstin, row.invoice_series)].append(row)
        for (_gstin, _series), items in groups.items():
            if len(items) < 3:
                continue
            seqs = sorted({int(r.invoice_sequence_no) for r in items if r.invoice_sequence_no is not None})
            span = seqs[-1] - seqs[0] + 1
            # Safety guard: invoice numbers may contain embedded dates/years/ERP ids,
            # producing very large numeric spans. Gap detection is useful only for
            # compact running sequences.
            if span > max(5000, len(seqs) * 10):
                for row in items:
                    row.append_audit_note("GAP_DETECTION_SKIPPED: sequence span too large (ERP/date-embedded numbers).")
                continue
            seq_set = set(seqs)
            missing = [n for n in range(seqs[0], seqs[-1] + 1) if n not in seq_set]
            if missing:
                note = "Possible invoice sequence gap(s): " + ", ".join(str(n) for n in missing[:20])
                if len(missing) > 20:
                    note += f" + {len(missing) - 20} more"
                for row in items:
                    row.invoice_gap_note = note
                    row.detected_snapshot["invoice_gap_note"] = note
                    row.append_audit_note(note)

    def _annotate_supplier_anomalies(self, rows: List[InvoiceRow], cfg: AuditConfig) -> None:
        amounts: Dict[str, List[Decimal]] = defaultdict(list)
        rows_by_supplier: Dict[str, List[InvoiceRow]] = defaultdict(list)
        for row in rows:
            if row.include_in_totals and (row.gstin or row.supplier_name) and row.invoice_value > 0:
                key = supplier_group_key(row)
                amounts[key].append(row.invoice_value)
                rows_by_supplier[key].append(row)

        baselines: Dict[str, Decimal] = {}
        for supplier, values in amounts.items():
            if len(values) >= 5:
                baselines[supplier] = Decimal(str(median([float(v) for v in values]))).quantize(Decimal("0.01"))
            elif len(values) >= 2:
                baselines[supplier] = (sum(values) / len(values)).quantize(Decimal("0.01"))
            else:
                for row in rows_by_supplier.get(supplier, []):
                    row.append_audit_note("ANOMALY_INSUFFICIENT_HISTORY: only 1 invoice; anomaly detection skipped.")

        for row in rows:
            baseline = baselines.get(supplier_group_key(row))
            if baseline and baseline > 0 and row.invoice_value > baseline * cfg.supplier_anomaly_multiplier:
                row.anomaly_note = f"Invoice value is more than {cfg.supplier_anomaly_multiplier}x supplier baseline ({baseline})."
                row.append_audit_note(row.anomaly_note)
                row.detected_snapshot["anomaly_note"] = row.anomaly_note

    def build_summary(self, rows: List[InvoiceRow], files_processed: int, sheets_processed: int) -> AuditSummary:
        summary = AuditSummary(files_processed=files_processed, sheets_processed=sheets_processed)
        summary.raw_rows_read = len(rows)
        summary.classified_rows = len(rows)
        official_rows = [row for row in rows if is_official_invoice_detail_row(row)]
        summary.official_invoice_rows = len(official_rows)
        for row in rows:
            if row.audit_status == "VALID":
                summary.valid_rows += 1
            if row.review_required:
                summary.review_required_rows += 1
            if row.audit_status.startswith("SKIPPED") or row.audit_status.startswith("ERROR"):
                summary.skipped_rows += 1
            if row.audit_status == "DUPLICATE_EXCLUDED":
                summary.duplicate_rows += 1
            if row.include_in_totals:
                summary.final_approved_rows += 1
                if row.audit_status != "VALID":
                    summary.accepted_warning_rows += 1
                summary.approved_invoice_value += row.invoice_value
                summary.approved_taxable_value += row.taxable_value
                summary.approved_igst += row.igst
                summary.approved_cgst += row.cgst
                summary.approved_sgst += row.sgst
                summary.approved_cess += row.cess
            elif row.review_required:
                summary.review_invoice_value += row.invoice_value
            else:
                summary.excluded_invoice_value += row.invoice_value
            if row.invoice_value:
                summary.raw_detected_invoice_value += row.invoice_value
            if row.mismatch_reason and row.mismatch_reason not in {"BALANCED_OR_ROUNDING", "MINOR_ROUNDING_OR_DECIMAL_ISSUE", "CREDIT_NOTE_BALANCED", "CREDIT_NOTE_ZERO_RATED"}:
                summary.gst_mismatch_rows += 1
            if row.audit_severity == "HIGH":
                summary.high_severity_rows += 1
            if row.audit_severity == "CRITICAL":
                summary.critical_rows += 1

        summary.approved_total_gst = summary.approved_igst + summary.approved_cgst + summary.approved_sgst + summary.approved_cess
        # Dashboard semantics are intentionally split:
        #   unique_* = approved suppliers/GSTINs that contribute to official totals.
        #   detected_unique_* = source-file suppliers/GSTINs found in official invoice/detail rows,
        #                       including rows awaiting review or excluded as duplicates.
        # This prevents v9.4's approved-only counter from hiding source-file coverage.
        approved_identity_rows = (
            [r for r in official_rows if r.include_in_totals]
            or [r for r in rows if (r.gstin or r.supplier_name) and r.include_in_totals]
        )
        detected_identity_rows = (
            [r for r in official_rows if (r.gstin or r.supplier_name)]
            or [r for r in rows if (r.gstin or r.supplier_name)]
        )
        summary.unique_suppliers = len({(r.gstin or supplier_group_key(r)) for r in approved_identity_rows if (r.gstin or r.supplier_name)})
        summary.unique_gstins = len({r.gstin for r in approved_identity_rows if r.gstin})
        summary.detected_unique_suppliers = len({(r.gstin or supplier_group_key(r)) for r in detected_identity_rows if (r.gstin or r.supplier_name)})
        summary.detected_unique_gstins = len({r.gstin for r in detected_identity_rows if r.gstin})
        summary.row_coverage_status = "MATCHED" if summary.raw_rows_read == summary.classified_rows else "FAILED"
        lhs = summary.raw_detected_invoice_value.quantize(Decimal("0.01"))
        rhs = (summary.approved_invoice_value + summary.review_invoice_value + summary.excluded_invoice_value).quantize(Decimal("0.01"))
        summary.amount_reconciliation_status = "MATCHED" if lhs == rhs else "FAILED"
        if summary.row_coverage_status == "MATCHED" and summary.amount_reconciliation_status == "MATCHED" and summary.review_required_rows == 0:
            summary.final_status = "FULLY_VERIFIED"
        elif summary.row_coverage_status == "MATCHED" and summary.amount_reconciliation_status == "MATCHED":
            summary.final_status = "BALANCED_BUT_REVIEW_REQUIRED"
        else:
            summary.final_status = "RECONCILIATION_FAILED"
        return summary

    def group_totals(self, rows: List[InvoiceRow], attr: str) -> Dict[str, Decimal]:
        totals: Dict[str, Decimal] = {}
        for row in rows:
            if not row.include_in_totals:
                continue
            key = getattr(row, attr) or "UNKNOWN"
            totals[key] = totals.get(key, Decimal("0.00")) + row.invoice_value
        return dict(sorted(totals.items(), key=lambda item: item[1], reverse=True))

    def supplier_totals(self, rows: List[InvoiceRow]) -> Dict[str, Decimal]:
        totals: Dict[str, Decimal] = {}
        for row in rows:
            if not row.include_in_totals:
                continue
            key = supplier_group_key(row)
            totals[key] = totals.get(key, Decimal("0.00")) + row.invoice_value
        return dict(sorted(totals.items(), key=lambda item: item[1], reverse=True))

    def month_totals(self, rows: List[InvoiceRow]) -> Dict[str, Decimal]:
        totals: Dict[str, Decimal] = {}
        for row in rows:
            if not row.include_in_totals:
                continue
            if row.invoice_date:
                key = row.invoice_date.strftime("%b %Y")
            elif row.period:
                key = row.period
            else:
                key = "UNKNOWN"
            totals[key] = totals.get(key, Decimal("0.00")) + row.invoice_value
        return dict(sorted(totals.items()))

    # Backwards-compatible aliases used by older scripts/tests.
    _build_summary = build_summary
    _group_totals = group_totals
    _month_totals = month_totals
