from __future__ import annotations

import hashlib
import re
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Iterable, Sequence

import pandas as pd
from openpyxl import load_workbook

GSTIN_RE = re.compile(r"\b\d{2}[A-Z]{5}\d{4}[A-Z][1-9A-Z]Z[0-9A-Z]\b", re.I)
PERIOD_RE = re.compile(r"\b(?P<month>jan(?:uary)?|feb(?:ruary)?|mar(?:ch)?|apr(?:il)?|may|jun(?:e)?|jul(?:y)?|aug(?:ust)?|sep(?:t(?:ember)?)?|oct(?:ober)?|nov(?:ember)?|dec(?:ember)?)\s*['\- ]?\s*(?P<year>\d{2}|\d{4})\b", re.I)

MONTHS = {
    "jan": 1, "january": 1,
    "feb": 2, "february": 2,
    "mar": 3, "march": 3,
    "apr": 4, "april": 4,
    "may": 5,
    "jun": 6, "june": 6,
    "jul": 7, "july": 7,
    "aug": 8, "august": 8,
    "sep": 9, "sept": 9, "september": 9,
    "oct": 10, "october": 10,
    "nov": 11, "november": 11,
    "dec": 12, "december": 12,
}

FULL_WORKBOOK_MARKERS = {
    "readme", "read me", "alltables", "all tables", "itcavailable", "itc available",
    "itcnotavailable", "itc not available", "itcreversal", "itc reversal", "itcrejected", "itc rejected",
    "b2ba", "b2bcdnr", "b2b cdnr", "b2bcdnra", "isd", "impg", "impgsez",
}

B2B_SHEET_NORMALIZED = {"b2b"}


@dataclass(frozen=True)
class ImportFileProfile:
    path: str
    file_name: str
    file_type: str
    period: str
    period_source: str
    sheet_count: int
    b2b_sheet_names: tuple[str, ...]
    b2b_row_count: int
    b2b_invoice_value: Decimal
    b2b_gst_value: Decimal
    b2b_hash: str
    confidence: int
    issue: str = ""

    @property
    def is_full_workbook(self) -> bool:
        return self.file_type == "FULL_GSTR2B_WORKBOOK"

    @property
    def usable(self) -> bool:
        return bool(self.period and self.b2b_hash and self.b2b_row_count > 0)


@dataclass(frozen=True)
class DuplicatePeriodGroup:
    period: str
    recommended_path: str
    recommended_file: str
    duplicate_paths: tuple[str, ...]
    duplicate_files: tuple[str, ...]
    exact_duplicate: bool
    reason: str
    prevented_rows: int
    prevented_invoice_value: Decimal
    prevented_gst_value: Decimal


@dataclass(frozen=True)
class ImportSafetyReport:
    profiles: tuple[ImportFileProfile, ...]
    duplicate_groups: tuple[DuplicatePeriodGroup, ...]
    recommended_paths: tuple[str, ...]
    excluded_duplicate_paths: tuple[str, ...]
    unresolved_duplicate_periods: tuple[str, ...]
    missing_periods: tuple[str, ...]
    uploaded_file_count: int
    unique_period_count: int
    selected_file_count: int
    duplicate_file_count: int
    blocked: bool
    status: str
    score: int
    message: str

    def recommended_file_names(self) -> list[str]:
        return [Path(p).name for p in self.recommended_paths]

    def excluded_file_names(self) -> list[str]:
        return [Path(p).name for p in self.excluded_duplicate_paths]

    def summary_lines(self) -> list[str]:
        lines = [
            f"Uploaded files: {self.uploaded_file_count}",
            f"Unique GST periods: {self.unique_period_count}",
            f"Selected for audit: {self.selected_file_count}",
            f"Duplicate files excluded: {self.duplicate_file_count}",
        ]
        if self.missing_periods:
            lines.append("Missing periods: " + ", ".join(self.missing_periods))
        if self.unresolved_duplicate_periods:
            lines.append("Unresolved duplicate periods: " + ", ".join(self.unresolved_duplicate_periods))
        lines.append(self.message)
        return lines


def normalize_sheet_name(name: object) -> str:
    return re.sub(r"[^a-z0-9]", "", str(name or "").strip().lower())


def _normalize_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    text = re.sub(r"\s+", " ", text)
    return text


def _normalize_money(value: object) -> str:
    text = _normalize_text(value).replace(",", "").replace("₹", "")
    if not text:
        return "0.00"
    try:
        return str(Decimal(text).quantize(Decimal("0.01")))
    except (InvalidOperation, ValueError):
        return text.upper()


def _decimal(value: object) -> Decimal:
    text = _normalize_money(value)
    try:
        return Decimal(text)
    except (InvalidOperation, ValueError):
        return Decimal("0.00")


def _period_from_match(match: re.Match[str]) -> str:
    month_text = match.group("month").lower()
    month = MONTHS[month_text[:3] if month_text[:3] in MONTHS else month_text]
    year_text = match.group("year")
    year = int(year_text)
    if year < 100:
        year += 2000
    return f"{year:04d}-{month:02d}"


def period_label(period: str) -> str:
    if not period or not re.match(r"^\d{4}-\d{2}$", period):
        return period or "UNKNOWN"
    year, month = period.split("-")
    names = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    return f"{names[int(month)-1]} {year}"


def detect_period_from_filename(path: Path) -> tuple[str, str]:
    match = PERIOD_RE.search(path.stem)
    if match:
        return _period_from_match(match), "filename"
    return "", "unknown"


def _period_from_tax_period_and_fy(month_text: str, financial_year: str) -> str:
    month_key = str(month_text or "").strip().lower()
    if not month_key:
        return ""
    month = MONTHS.get(month_key) or MONTHS.get(month_key[:3])
    if not month:
        return ""
    fy_match = re.search(r"(\d{4})\s*[-/]\s*(\d{2}|\d{4})", str(financial_year or ""))
    if not fy_match:
        return ""
    start_year = int(fy_match.group(1))
    year = start_year if month >= 4 else start_year + 1
    return f"{year:04d}-{month:02d}"


def _detect_period_from_rows(path: Path, workbook, b2b_sheet_names: Sequence[str]) -> tuple[str, str]:
    # Full GSTR-2B workbooks expose the real tax period in Read me as
    # Financial Year + Tax Period. This is more reliable than B2B row filing
    # period columns, which may point to supplier filing month rather than the
    # recipient's selected GSTR-2B period.
    for ws in workbook.worksheets:
        norm = normalize_sheet_name(ws.title)
        if not (norm in {"readme", "readme1", "summary"} or "readme" in norm):
            continue
        row_limit = min(ws.max_row or 1, 30)
        fy = ""
        tax_period = ""
        for row in ws.iter_rows(min_row=1, max_row=row_limit, max_col=4, values_only=True):
            cells = [_normalize_text(v) for v in row]
            for idx, cell in enumerate(cells):
                low = cell.lower()
                nxt = cells[idx + 1] if idx + 1 < len(cells) else ""
                if low == "financial year" and nxt:
                    fy = nxt
                elif low == "tax period" and nxt:
                    tax_period = nxt
        period = _period_from_tax_period_and_fy(tax_period, fy)
        if period:
            return period, f"sheet:{ws.title}"

    # B2B-only files generally do not contain reliable workbook metadata. In
    # that case, filename is safer than the GSTR filing-period column.
    filename_period, filename_source = detect_period_from_filename(path)
    if filename_period:
        return filename_period, filename_source

    # Last resort only: scan visible worksheet text.
    candidate_sheets = [workbook[s] for s in b2b_sheet_names]
    for ws in candidate_sheets:
        row_limit = min(ws.max_row or 1, 20)
        col_limit = min(ws.max_column or 1, 20)
        for row in ws.iter_rows(min_row=1, max_row=row_limit, max_col=col_limit, values_only=True):
            joined = " ".join(_normalize_text(v) for v in row if _normalize_text(v))
            match = PERIOD_RE.search(joined)
            if match:
                return _period_from_match(match), f"sheet:{ws.title}"
    return "", "unknown"


def _classify_workbook(sheet_names: Sequence[str]) -> str:
    normalized = {normalize_sheet_name(s) for s in sheet_names}
    if len(sheet_names) > 1 and (normalized & {normalize_sheet_name(m) for m in FULL_WORKBOOK_MARKERS}):
        return "FULL_GSTR2B_WORKBOOK"
    if normalized == {"b2b"} or (len(sheet_names) == 1 and "b2b" in normalized):
        return "B2B_ONLY_SHEET"
    if "b2b" in normalized and len(sheet_names) > 1:
        return "FULL_GSTR2B_WORKBOOK"
    return "UNKNOWN_EXCEL"


def _find_b2b_sheets(sheet_names: Sequence[str]) -> tuple[str, ...]:
    result: list[str] = []
    for name in sheet_names:
        norm = normalize_sheet_name(name)
        if norm == "b2b":
            result.append(name)
    if not result:
        for name in sheet_names:
            if normalize_sheet_name(name).startswith("b2b"):
                result.append(name)
    return tuple(result)


def _find_header_indexes(values: Sequence[object]) -> dict[str, int]:
    lowered = [_normalize_text(v).lower() for v in values]
    indexes: dict[str, int] = {}
    for i, text in enumerate(lowered):
        if "gstin" in text and "supplier" in text:
            indexes["gstin"] = i
        elif "trade" in text or "legal" in text:
            indexes.setdefault("supplier", i)
        elif "taxable" in text:
            indexes["taxable"] = i
        elif "invoice value" in text:
            indexes["invoice_value"] = i
        elif "integrated tax" in text:
            indexes["igst"] = i
        elif "central tax" in text:
            indexes["cgst"] = i
        elif "state" in text and "tax" in text:
            indexes["sgst"] = i
    # GSTR-2B has multi-row header: invoice number/date/value are on row below.
    for i, text in enumerate(lowered):
        if "invoice number" in text:
            indexes["invoice_no"] = i
        elif "invoice date" in text:
            indexes["invoice_date"] = i
        elif "invoice value" in text:
            indexes["invoice_value"] = i
    return indexes


def _extract_b2b_fingerprint(workbook, b2b_sheet_names: Sequence[str]) -> tuple[int, Decimal, Decimal, str]:
    normalized_records: list[str] = []
    invoice_value_total = Decimal("0.00")
    gst_value_total = Decimal("0.00")
    for sheet_name in b2b_sheet_names:
        ws = workbook[sheet_name]
        header_indexes: dict[str, int] = {}
        rows_cache: list[tuple[int, tuple[object, ...]]] = []
        for row_no, row in enumerate(ws.iter_rows(values_only=True), start=1):
            row_tuple = tuple(row)
            rows_cache.append((row_no, row_tuple))
            row_text = " ".join(_normalize_text(v).lower() for v in row_tuple)
            if "gstin of supplier" in row_text or "invoice number" in row_text:
                header_indexes.update(_find_header_indexes(row_tuple))
                # Merge with next row if it contains invoice/date/value labels.
                continue
            if header_indexes and "invoice_no" not in header_indexes:
                header_indexes.update(_find_header_indexes(row_tuple))
            if row_no > 20 and header_indexes:
                break
        # Process every row after header scan; B2B rows contain supplier GSTIN.
        for row_no, row in rows_cache + list(enumerate(ws.iter_rows(min_row=max(len(rows_cache) + 1, 1), values_only=True), start=len(rows_cache) + 1)):
            values = list(row)
            row_joined = " ".join(_normalize_text(v) for v in values)
            match = GSTIN_RE.search(row_joined.upper())
            if not match:
                continue
            gstin = match.group(0).upper()
            supplier = _normalize_text(values[1] if len(values) > 1 else "").upper()
            invoice_no = _normalize_text(values[2] if len(values) > 2 else "").upper()
            invoice_date = _normalize_text(values[4] if len(values) > 4 else "").upper()
            invoice_value = _normalize_money(values[5] if len(values) > 5 else "0")
            taxable = _normalize_money(values[8] if len(values) > 8 else "0")
            igst = _normalize_money(values[9] if len(values) > 9 else "0")
            cgst = _normalize_money(values[10] if len(values) > 10 else "0")
            sgst = _normalize_money(values[11] if len(values) > 11 else "0")
            cess = _normalize_money(values[12] if len(values) > 12 else "0") if len(values) > 12 else "0.00"
            # Avoid false rows: need a concrete invoice number and at least one amount.
            if not invoice_no or invoice_no in {"INVOICE NUMBER", "INVOICE DETAILS"}:
                continue
            if invoice_value == "0.00" and taxable == "0.00" and igst == "0.00" and cgst == "0.00" and sgst == "0.00":
                continue
            invoice_value_total += _decimal(invoice_value)
            gst_value_total += _decimal(igst) + _decimal(cgst) + _decimal(sgst) + _decimal(cess)
            normalized_records.append("|".join([
                gstin, supplier, invoice_no, invoice_date,
                taxable, igst, cgst, sgst, cess, invoice_value,
            ]))
    normalized_records.sort()
    payload = "\n".join(normalized_records).encode("utf-8", errors="replace")
    digest = hashlib.sha256(payload).hexdigest() if normalized_records else ""
    return len(normalized_records), invoice_value_total.quantize(Decimal("0.01")), gst_value_total.quantize(Decimal("0.01")), digest


def analyze_import_file(path: str | Path) -> ImportFileProfile:
    p = Path(path)
    try:
        wb = load_workbook(p, read_only=True, data_only=True)
        try:
            sheet_names = tuple(wb.sheetnames)
            file_type = _classify_workbook(sheet_names)
            b2b_sheets = _find_b2b_sheets(sheet_names)
            period, period_source = _detect_period_from_rows(p, wb, b2b_sheets)
            b2b_count, invoice_value, gst_value, b2b_hash = _extract_b2b_fingerprint(wb, b2b_sheets)
        finally:
            wb.close()
        confidence = 0
        if period:
            confidence += 30
        if file_type != "UNKNOWN_EXCEL":
            confidence += 25
        if b2b_hash:
            confidence += 35
        if b2b_count:
            confidence += 10
        issue = "" if confidence >= 80 else "Low confidence import profile; user review required."
        return ImportFileProfile(
            path=str(p), file_name=p.name, file_type=file_type, period=period, period_source=period_source,
            sheet_count=len(sheet_names), b2b_sheet_names=tuple(b2b_sheets), b2b_row_count=b2b_count,
            b2b_invoice_value=invoice_value, b2b_gst_value=gst_value, b2b_hash=b2b_hash,
            confidence=min(confidence, 100), issue=issue,
        )
    except Exception as exc:
        period, period_source = detect_period_from_filename(p)
        return ImportFileProfile(
            path=str(p), file_name=p.name, file_type="UNREADABLE", period=period, period_source=period_source,
            sheet_count=0, b2b_sheet_names=(), b2b_row_count=0, b2b_invoice_value=Decimal("0.00"),
            b2b_gst_value=Decimal("0.00"), b2b_hash="", confidence=0, issue=f"Unreadable file: {exc}",
        )


def _profile_priority(profile: ImportFileProfile) -> tuple[int, int, int, str]:
    type_score = {"FULL_GSTR2B_WORKBOOK": 3, "B2B_ONLY_SHEET": 2, "UNKNOWN_EXCEL": 1}.get(profile.file_type, 0)
    return (type_score, profile.confidence, profile.sheet_count, profile.file_name.lower())


def expected_fy_periods_from_profiles(profiles: Sequence[ImportFileProfile]) -> tuple[str, ...]:
    periods = sorted({p.period for p in profiles if re.match(r"^\d{4}-\d{2}$", p.period or "")})
    if not periods:
        return ()
    months = [int(p.split("-")[1]) for p in periods]
    years = [int(p.split("-")[0]) for p in periods]
    # GST FY in India is Apr-Mar; infer when data spans April to next year.
    if 4 in months:
        start_year = min(int(p.split("-")[0]) for p in periods if p.endswith("-04"))
        return tuple(f"{y:04d}-{m:02d}" for y, m in [(start_year if m >= 4 else start_year + 1, m) for m in range(4, 13)] + [(start_year + 1, m) for m in range(1, 4)])
    return ()


def analyze_import_set(file_paths: Iterable[str | Path], *, expected_periods: Sequence[str] | None = None) -> ImportSafetyReport:
    profiles = tuple(analyze_import_file(p) for p in file_paths)
    by_period: dict[str, list[ImportFileProfile]] = {}
    no_period: list[ImportFileProfile] = []
    for profile in profiles:
        if profile.period:
            by_period.setdefault(profile.period, []).append(profile)
        else:
            no_period.append(profile)

    recommended: list[str] = []
    excluded: list[str] = []
    duplicate_groups: list[DuplicatePeriodGroup] = []
    unresolved: list[str] = []

    for period, group in sorted(by_period.items()):
        sorted_group = sorted(group, key=_profile_priority, reverse=True)
        chosen = sorted_group[0]
        recommended.append(chosen.path)
        if len(sorted_group) == 1:
            continue
        chosen_hash = chosen.b2b_hash
        dup_profiles = sorted_group[1:]
        exact = all(p.b2b_hash and p.b2b_hash == chosen_hash for p in dup_profiles)
        if exact:
            excluded.extend(p.path for p in dup_profiles)
            duplicate_groups.append(DuplicatePeriodGroup(
                period=period,
                recommended_path=chosen.path,
                recommended_file=chosen.file_name,
                duplicate_paths=tuple(p.path for p in dup_profiles),
                duplicate_files=tuple(p.file_name for p in dup_profiles),
                exact_duplicate=True,
                reason="Same period and same B2B invoice hash. Full workbook is preferred when available.",
                prevented_rows=sum(p.b2b_row_count for p in dup_profiles),
                prevented_invoice_value=sum((p.b2b_invoice_value for p in dup_profiles), Decimal("0.00")).quantize(Decimal("0.01")),
                prevented_gst_value=sum((p.b2b_gst_value for p in dup_profiles), Decimal("0.00")).quantize(Decimal("0.01")),
            ))
        else:
            unresolved.append(period)
            duplicate_groups.append(DuplicatePeriodGroup(
                period=period,
                recommended_path=chosen.path,
                recommended_file=chosen.file_name,
                duplicate_paths=tuple(p.path for p in dup_profiles),
                duplicate_files=tuple(p.file_name for p in dup_profiles),
                exact_duplicate=False,
                reason="Same period but B2B hash differs. Manual selection required before audit.",
                prevented_rows=0,
                prevented_invoice_value=Decimal("0.00"),
                prevented_gst_value=Decimal("0.00"),
            ))

    # Non-period/unreadable files cannot be auto-selected safely. Keep them only when no period grouping exists.
    if no_period and not by_period:
        recommended.extend(p.path for p in no_period)

    expected = tuple(expected_periods or expected_fy_periods_from_profiles(profiles))
    present = set(by_period.keys())
    missing = tuple(period for period in expected if period not in present)
    blocked = bool(unresolved)
    duplicate_file_count = len(excluded)
    score = 100
    if blocked:
        score -= 40
    if no_period:
        score -= min(20, len(no_period) * 5)
    if missing:
        score -= min(12, len(missing) * 4)
    if duplicate_file_count:
        score -= 0  # exact duplicates are handled, not penalized after recommendation.
    score = max(0, min(100, score))
    if blocked:
        status = "BLOCKED_DUPLICATES_NEED_MANUAL_REVIEW"
        message = "Audit blocked: at least one period has conflicting duplicate files. Review duplicate groups first."
    elif duplicate_file_count:
        status = "SAFE_WITH_DUPLICATES_EXCLUDED"
        message = f"Safe recommended selection created: {len(recommended)} file(s) selected and {duplicate_file_count} duplicate file(s) excluded."
    else:
        status = "SAFE_TO_AUDIT"
        message = "No exact duplicate period files detected. Safe to audit selected files."
    if missing and not blocked:
        message += " Missing periods are reported but do not block partial-period audit."
    return ImportSafetyReport(
        profiles=profiles,
        duplicate_groups=tuple(duplicate_groups),
        recommended_paths=tuple(recommended),
        excluded_duplicate_paths=tuple(excluded),
        unresolved_duplicate_periods=tuple(unresolved),
        missing_periods=missing,
        uploaded_file_count=len(profiles),
        unique_period_count=len(by_period),
        selected_file_count=len(recommended),
        duplicate_file_count=duplicate_file_count,
        blocked=blocked,
        status=status,
        score=score,
        message=message,
    )


def import_safety_dataframes(report: ImportSafetyReport) -> dict[str, pd.DataFrame]:
    dashboard = pd.DataFrame([
        {"metric": "Uploaded files", "value": report.uploaded_file_count},
        {"metric": "Unique periods", "value": report.unique_period_count},
        {"metric": "Selected for audit", "value": report.selected_file_count},
        {"metric": "Duplicate files excluded", "value": report.duplicate_file_count},
        {"metric": "Missing periods", "value": ", ".join(period_label(p) for p in report.missing_periods) or "None"},
        {"metric": "Status", "value": report.status},
        {"metric": "Import safety score", "value": report.score},
        {"metric": "Message", "value": report.message},
    ])
    profiles = pd.DataFrame([
        {
            "file_name": p.file_name,
            "file_type": p.file_type,
            "period": period_label(p.period),
            "period_code": p.period,
            "period_source": p.period_source,
            "sheets": p.sheet_count,
            "b2b_sheets": ", ".join(p.b2b_sheet_names),
            "b2b_rows": p.b2b_row_count,
            "b2b_invoice_value": float(p.b2b_invoice_value),
            "b2b_gst_value": float(p.b2b_gst_value),
            "b2b_hash": p.b2b_hash,
            "confidence": p.confidence,
            "issue": p.issue,
            "path": p.path,
        }
        for p in report.profiles
    ])
    duplicate_groups = pd.DataFrame([
        {
            "period": period_label(g.period),
            "period_code": g.period,
            "recommended_file": g.recommended_file,
            "duplicate_files": ", ".join(g.duplicate_files),
            "exact_duplicate": g.exact_duplicate,
            "prevented_rows": g.prevented_rows,
            "prevented_invoice_value": float(g.prevented_invoice_value),
            "prevented_gst_value": float(g.prevented_gst_value),
            "reason": g.reason,
        }
        for g in report.duplicate_groups
    ])
    selected = pd.DataFrame([{"selected_file": Path(p).name, "path": p} for p in report.recommended_paths])
    excluded = pd.DataFrame([{"excluded_duplicate_file": Path(p).name, "path": p} for p in report.excluded_duplicate_paths])
    rules = pd.DataFrame([
        {"rule": "Detect period", "detail": "Read workbook metadata/B2B period first; fallback to filename."},
        {"rule": "Hash B2B rows", "detail": "Normalize GSTIN, supplier, invoice number/date, taxable/GST/invoice values, sort, SHA-256 hash."},
        {"rule": "Duplicate group", "detail": "Same period + same B2B hash = exact duplicate."},
        {"rule": "Preferred file", "detail": "Full GSTR-2B workbook beats B2B-only sheet."},
        {"rule": "Audit blocking", "detail": "Conflicting duplicate hashes block audit until manual selection."},
    ])
    return {
        "Dashboard": dashboard,
        "File Profiles": profiles,
        "Duplicate Groups": duplicate_groups,
        "Recommended Files": selected,
        "Excluded Duplicates": excluded,
        "Import Rules": rules,
    }


def export_import_safety_report(report: ImportSafetyReport, output_path: str | Path) -> Path:
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
        for sheet_name, df in import_safety_dataframes(report).items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            worksheet = writer.sheets[sheet_name]
            worksheet.freeze_panes(1, 0)
            for idx, column in enumerate(df.columns):
                width = min(max(len(str(column)) + 2, 14), 55)
                if not df.empty:
                    width = min(max(width, int(df[column].astype(str).str.len().quantile(0.9)) + 2), 55)
                worksheet.set_column(idx, idx, width)
    return output
